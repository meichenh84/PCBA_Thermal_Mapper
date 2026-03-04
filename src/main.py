"""
=============================================================================
PCBA 熱力圖溫度點位自動識別工具 - 主程式 (main.py)
=============================================================================
本程式為「PCBA 熱力圖溫度點位自動識別工具」的核心主程式，負責建立整個桌面應用
程式的圖形化使用者介面 (GUI)，並整合資料夾管理、圖像對齊、溫度過濾、溫度編輯、
匯出報告等功能。

功能概述：
  1. 資料夾管理：掃描資料夾中的熱力圖、Layout圖、溫度數據、Layout數據並分類
  2. 圖像對齊：在熱力圖與Layout圖上標記對齊點，計算座標變換矩陣
  3. 溫度過濾：根據Layout元器件位置與溫度數據，自動識別高溫區域
  4. 溫度編輯：透過EditorCanvas對話框手動調整/新增/刪除標記框
  5. 匯出報告：將識別結果匯出為Excel報表與標記圖片

介面結構：
  上方：工具列 [資料夾Tab] [對齊圖像] [溫度過濾] [匯出] [設定]
  左側：資料夾檔案樹狀列表 (folder_tree)
  中間左：canvasA（熱力圖顯示區）
  中間右：canvasB（Layout圖顯示區）
  下方：對齊前/後圖像預覽按鈕、清除對齊點按鈕

主要類別：
  - ResizableImagesApp：應用程式主類別，管理所有 UI 元件與業務邏輯

版本：V20251011
=============================================================================
"""
import sys
import os
# 設定標準輸出編碼為 UTF-8，避免中文輸出時產生亂碼
os.environ['PYTHONIOENCODING'] = 'utf-8'
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# ===== 標準函式庫與第三方套件匯入 =====
import tkinter as tk                          # Tkinter GUI 框架
from tkinter import Canvas, filedialog, messagebox, simpledialog, ttk  # Tkinter 子模組（畫布、檔案對話框、訊息框、樹狀列表）
from PIL import Image, ImageTk                # Pillow 圖像處理庫（讀取/縮放/轉換圖片）
from magnifier import ImageMagnifier          # 自訂放大鏡元件（對齊模式下輔助精確打點）
from load_tempA import TempLoader             # 溫度數據載入器（讀取 CSV/XLSX 溫度矩陣）
from toast import show_toast                  # 浮動通知訊息元件（成功/警告/錯誤提示）
import cv2                                    # OpenCV 圖像處理（色彩轉換、圖像混合、仿射變換）
import numpy as np                            # NumPy 數值運算（矩陣計算、座標變換）
import pandas as pd                           # Pandas 數據分析（讀取 Excel/CSV 檔案）
import openpyxl                               # Excel 檔案操作（匯出報告）
import json                                   # JSON 序列化/反序列化（儲存/讀取對齊點資料）
import threading                              # 多執行緒支援（非同步載入大型檔案）
import time                                   # 時間相關工具
import math                                   # 數學運算（旋轉角度計算）
import argparse                               # 命令列參數解析
from datetime import datetime                 # 日期時間處理

# ===== 自訂模組匯入 =====
from dialog_template import TemplateDialog    # 溫度過濾參數設定對話框
from dialog_setting import SettingDialog      # 應用程式設定對話框
from constants import Constants               # 全域常數定義（預設路徑等）
from point_transformer import PointTransformer  # 座標變換器（A圖↔B圖座標轉換，支援仿射/透視變換）
from config import GlobalConfig               # 全域配置管理器（儲存/讀取使用者偏好設定）

# UI 樣式常數定義 —— 匯入 UIStyle 以保持全應用程式的視覺樣式統一
try:
    from .ui_style import UIStyle
except ImportError:
    from ui_style import UIStyle
from circle_ring_draw import draw_points_circle_ring_text, draw_points_circle_ring  # 繪製帶編號的圓環對齊點標記
from recognize_circle import detect_A_circles, detect_B_circles, find_circle_containing_point  # 圓形區域偵測（輔助打點吸附）
from draw_rect import draw_triangle_and_text, draw_canvas_item, update_canvas_item, draw_numpy_image_item  # 繪製矩形溫度標記框
from editor_canvas import EditorCanvas        # 溫度標記編輯畫布（獨立對話框，可調整/新增/刪除標記框）
from datetime import datetime
import csv                                    # CSV 檔案操作（匯出日誌）
import copy                                   # 深層複製工具（複製編輯日誌範本）
from layout_temperature_query_optimized import LayoutTemperatureQueryOptimized  # Layout 溫度查詢最佳化模組
from temperature_config_manager import TemperatureConfigManager  # 溫度配置管理器（每個資料夾獨立配置）
from folder_scanner import scan_folder, validate_layout_data     # 資料夾檔案自動分類與驗證
from dialog_export_report import ExportReportDialog              # 測試報告 Sheet 名稱選擇對話框
import tempfile                                                  # 暫存檔案（寫入測試報告用）
from openpyxl.drawing.image import Image as XlImage              # Excel 插入圖片


# ===== 編輯日誌預設範本 =====
# 記錄每次「溫度過濾 -> 編輯 -> 匯出」的操作統計資訊
# 結構：{ 欄位名: [中文說明, 初始值] }
DEFAULT_EDIT_LOG = {
    "export_time": ["生成时间", ""],                 # 匯出時間戳記
    "origin_mark": ["自动生成外框数量", 0],           # 自動識別產生的標記框數量
    "final_mark": ["最终导出外框数量", 0],            # 最終匯出時的標記框數量
    "add_new_mark": ["新增外框数量（手动增加导出时没有被删除）", 0],       # 使用者手動新增的標記框數量
    "delete_origin_mark": ["删除外框数量（自动生成的外框被删除）", 0],     # 使用者刪除的自動標記框數量
    "modify_origin_mark": ["调整外框数量（自动生成的外框被调整)", set()],  # 使用者調整過的標記框集合
}

def cv2_imread_unicode(image_path):
    """
    讀取含有中文（Unicode）路徑的圖片檔案。

    解決 OpenCV 在 Windows 上無法直接讀取含中文字元路徑的問題，
    先用 NumPy 將檔案讀為位元組陣列，再用 cv2.imdecode 解碼為圖片。

    參數：
        image_path (str): 圖片檔案的完整路徑（可包含中文字元）

    回傳：
        numpy.ndarray 或 None: 成功時回傳 BGR 格式的圖片陣列，失敗時回傳 None
    """
    try:
        # 使用 NumPy 讀取檔案為位元組陣列，繞過 OpenCV 的路徑編碼限制
        img_array = np.fromfile(image_path, dtype=np.uint8)
        # 將位元組陣列解碼為彩色圖片（BGR 格式）
        image = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
        return image
    except Exception as e:
        print(f"cv2_imread_unicode error: {e}")
        return None

class ResizableImagesApp:
    """
    PCBA 熱力圖溫度點位自動識別 - 主應用程式類別。

    本類別為整個應用程式的核心，負責管理所有 UI 元件、事件處理與業務邏輯。
    採用 Tkinter 的 Grid 佈局管理器來安排介面元素。

    核心功能：
        1. 熱力圖 (imageA) 與 Layout 圖 (imageB) 的座標映射（透過 PointTransformer）
        2. 溫度數據的智慧查詢和過濾（透過 LayoutTemperatureQueryOptimized）
        3. 元器件邊界的自動識別（根據 Layout 數據中的 RefDes、座標、尺寸）
        4. 溫度數據的視覺化顯示（矩形標記框繪製在 Canvas 上）
        5. 溫度標記的手動編輯（透過 EditorCanvas 對話框）
        6. 匯出 Excel 報表和標記圖片

    主要屬性：
        canvasA (tk.Canvas): 左側畫布，顯示熱力圖
        canvasB (tk.Canvas): 右側畫布，顯示 Layout 圖
        points_A (list): 熱力圖上的對齊點座標列表 [[x1,y1], [x2,y2], ...]
        points_B (list): Layout 圖上的對齊點座標列表
        mark_rect_A (list): 熱力圖上的溫度標記矩形框列表
        mark_rect_B (list): Layout 圖上的溫度標記矩形框列表
        point_transformer (PointTransformer): 座標變換器（A圖↔B圖）
        folder_tree (ttk.Treeview): 左側資料夾檔案樹狀列表
        current_folder_path (str): 當前資料夾路徑
        current_files (dict): 當前使用的各類型檔案 {heat, layout, heatTemp, layoutXY, layoutLWT}
    """
    def __init__(self, root):
        """
        初始化主應用程式。

        建立主視窗、初始化所有狀態變數、建構 UI 介面、綁定事件，
        並在延遲 100ms 後自動載入上次使用的資料夾。

        參數：
            root (tk.Tk): Tkinter 根視窗物件
        """
        print("V20251011")
        
        # 執行緒鎖，用於保護多執行緒環境下的共享資源
        self.lock = threading.Lock()
        self.root = root  # Tkinter 根視窗物件的參考
        
        # 設定主視窗屬性
        self.root.title("Thermal温度点位自动识别")  # 視窗標題
        self.root.minsize(width=400, height=500)       # 最小視窗尺寸
        self.root.geometry("1200x600")                  # 初始視窗尺寸
        
        # 畫布尺寸初始化（用於偵測視窗大小是否變化，避免重複渲染）
        self.canvasA_width = 1   # canvasA（熱力圖畫布）的當前寬度
        self.canvasA_height = 1  # canvasA（熱力圖畫布）的當前高度
        
        # 圖像對齊狀態控制
        self.is_aligning = False  # 是否處於「對齊模式」（True=打點中，False=一般瀏覽模式）

        # 矩形對齊模式狀態
        self.is_rect_aligning = False       # 矩形對齊模式
        self.rect_corners = None            # [TL, TR, BR, BL] 原始圖座標
        self.rect_drag_start = None         # 拖曳起點（canvas 座標）
        self.rect_dragging_corner = None    # 正在拖曳的角索引 (0-3)
        self.rect_canvas_ids = []           # overlay 的 canvas item IDs
        self.alignment_type = 'multi_point' # 'multi_point' 或 'rect'
        
        # 配置管理器
        self.config = GlobalConfig()    # 全域配置管理器（儲存視窗偏好、上次資料夾路徑等）
        self.temp_config = None         # 溫度配置管理器，每個資料夾有獨立的 temperature_config.json
        
        # 放大鏡元件（對齊模式下啟用，輔助使用者精確標記對齊點）
        self.canvasA_magnifier = None  # 熱力圖畫布的放大鏡實例
        self.canvasB_magnifier = None  # Layout 圖畫布的放大鏡實例
        
        # 圖像對齊相關數據（原始圖片座標，非畫布座標）
        self.points_A = []  # 熱力圖上的對齊點座標列表 [[x1,y1], [x2,y2], ...]
        self.points_B = []  # Layout 圖上的對齊點座標列表（與 points_A 一一對應）
        
        # 自動識別的圓形區域（用於打點時的吸附功能）
        self.recognize_circle_A = []  # 熱力圖上識別的圓形區域
        self.recognize_circle_B = []  # Layout 圖上識別的圓形區域
        
        # 溫度標記矩形框（溫度過濾後產生的高溫元器件標記）
        self.mark_rect_A = []  # 熱力圖上的溫度標記矩形框列表（每項為一個 dict）
        
        # 畫布背景圖像 ID（Canvas.create_image 回傳的 item ID）
        self.bg_imageA_id = None  # 熱力圖在 canvasA 上的圖片物件 ID
        self.bg_imageB_id = None  # Layout 圖在 canvasB 上的圖片物件 ID
        
        # 座標變換器（用於熱力圖與 Layout 圖之間的座標互相轉換）
        self.point_transformer = None  # PointTransformer 實例，對齊完成後建立
        
        # 圖像數據（原始尺寸的 PIL Image 物件）
        self.imageA = None  # 熱力圖的原始圖像數據 (PIL.Image)
        self.imageB = None  # Layout 圖的原始圖像數據 (PIL.Image)
        
        # 狀態旗標
        self.pont_marked = False  # 對齊點是否有被新增/刪除過（用於判斷是否需要清除舊標記框）
        self.edit_log = None      # 編輯日誌記錄（追蹤本次操作的新增/刪除/修改數量）
        
        # 資料夾選擇相關變數
        self.current_folder_path = None  # 當前工作資料夾的完整路徑
        self.folder_files = {"heat": [], "layout": [], "heatTemp": [], "layoutXY": [], "layoutLWT": [], "testReport": []}  # 資料夾中各分類的檔案列表
        self.current_temp_file_path = None  # 當前使用的溫度數據檔案完整路徑
        self.current_files = {"heat": None, "layout": None, "heatTemp": None, "layoutXY": None, "layoutLWT": None, "testReport": None}  # 各分類中當前選用的檔案名稱
        
        # Layout 數據相關變數
        self.layout_data = None  # 儲存解析後的 Layout 元器件數據（list of dict，含 RefDes、座標、尺寸等）
        self._xlsx_columns_cache = {}  # xlsx 欄位快取，供 _validate_layout_data() 使用
        self._category_warnings = set()  # 有驗證問題的分類名稱集合（用於顯示 ❗ 標記）
        self._tree_tooltip = None  # Treeview tooltip 視窗
        self._tree_tooltip_item = None  # 當前 tooltip 對應的 item

        # 對話框實例（單例模式，避免重複開啟）
        self.setting_dialog = None  # 設定對話框實例 (SettingDialog)
        self.editor_canvas = None   # 溫度標記編輯畫布實例 (EditorCanvas)

        # self.save_log_file()

        self.init_UI_flow(root)  # 建構完整的 UI 介面佈局

        # 初始化時顯示圖片（目前已改為由資料夾自動載入）
        # self.update_images()

        # 綁定視窗大小變化事件，觸發圖片重新縮放
        self.root.bind("<Configure>", self.on_resize)

        # 控制更新頻率的延遲計時器 ID（防止視窗縮放時頻繁觸發重繪）
        self.resize_after = None
        # self.root.after(100, self.init_magnifier)  # 延迟100毫秒更新
        # self.background_opt()
        self.root.after(100, self.background_opt)  # 延遲 100ms 後執行背景初始化（載入上次資料夾等）

    def background_opt(self):
        """背景初始化作業，在主視窗建立後延遲執行，避免阻塞 UI 啟動。"""
        # 載入上次使用的資料夾路徑並自動恢復工作狀態
        self.load_last_folder_path()
        
    def load_last_folder_path(self):
        """從全域配置中載入上次使用的資料夾路徑，並自動恢復工作狀態。

        若上次的資料夾路徑仍然存在，則自動掃描檔案、更新檔案樹、
        載入圖片與對齊點數據，讓使用者無需重新選擇資料夾。
        """
        last_path = self.config.get("last_folder_path")
        if last_path and os.path.exists(last_path):
            print(f"启动时自动加载上次使用的文件夹: {last_path}")
            
            # 保存当前文件夹的文件选择
            self.save_current_files_to_config()
            
            # 清空旧的数据
            self.clear_old_data()
            
            self.current_folder_path = last_path
            
            # 初始化温度配置管理器（重要：必须在scan_folder_files之前）
            self.temp_config = TemperatureConfigManager(last_path)
            print(f"已初始化TemperatureConfigManager，配置文件路径: {last_path}/config/temperature_config.json")
            
            self.scan_folder_files()
            self.update_folder_display()
            self.update_folder_path_label()
            
            # 自动加载图片和点位数据
            self.auto_load_images()
            
            # 更新按钮文字
            folder_name = os.path.basename(last_path)
            self.folder_control_button.config(text=f"隐藏文件夹Tab")
            
            print(f"启动时文件夹自动加载完成: {folder_name}")
        else:
            print("没有找到上次使用的文件夹路径或文件夹不存在")
    
    def save_folder_path(self):
        """將當前資料夾路徑儲存到全域配置檔案中，以便下次啟動時自動載入。"""
        if self.current_folder_path:
            self.config.set("last_folder_path", self.current_folder_path)
            self.config.save_to_json()
    
    def save_current_files_to_config(self):
        """將當前各分類中選用的檔案名稱儲存到 temperature_config.json 配置檔中。

        在切換資料夾或關閉程式前呼叫，確保下次開啟時能恢復檔案選擇狀態。
        """
        if self.current_folder_path:
            # 保存当前选择的文件到temperature_config.json
            if self.temp_config:
                self.temp_config.set_file_path("current_heat_file", self.current_files.get("heat"))
                self.temp_config.set_file_path("current_pcb_file", self.current_files.get("heat"))
                self.temp_config.set_file_path("current_layout_file", self.current_files.get("layout"))
                self.temp_config.set_file_path("current_temp_file", self.current_files.get("heatTemp"))
                self.temp_config.set_file_path("current_layout_xy_file", self.current_files.get("layoutXY"))
                self.temp_config.set_file_path("current_layout_lwt_file", self.current_files.get("layoutLWT"))
                print(f"已保存当前文件选择到temperature_config.json: 热力图={self.current_files.get('heat')}, Layout图={self.current_files.get('layout')}, 温度数据={self.current_files.get('heatTemp')}, 元器件座標={self.current_files.get('layoutXY')}, 元器件尺寸={self.current_files.get('layoutLWT')}")
            else:
                print("temp_config未初始化，无法保存文件路径")
    
    def update_temp_config_files(self):
        """同步更新溫度配置管理器中的當前檔案資訊。

        將 self.current_files 中的各分類檔案路徑寫入 temperature_config.json，
        供溫度過濾對話框等其他模組讀取。
        """
        print(f"update_temp_config_files: 开始更新文件信息")
        print(f"update_temp_config_files: temp_config存在: {self.temp_config is not None}")
        print(f"update_temp_config_files: current_folder_path: {self.current_folder_path}")
        print(f"update_temp_config_files: current_files: {self.current_files}")
        
        if self.temp_config and self.current_folder_path:
            # 使用新的文件路径管理方法
            self.temp_config.set_file_path("current_heat_file", self.current_files.get("heat"))
            self.temp_config.set_file_path("current_pcb_file", self.current_files.get("heat"))
            self.temp_config.set_file_path("current_temp_file", self.current_files.get("heatTemp"))
            self.temp_config.set_file_path("current_layout_file", self.current_files.get("layout"))
            self.temp_config.set_file_path("current_layout_xy_file", self.current_files.get("layoutXY"))
            self.temp_config.set_file_path("current_layout_lwt_file", self.current_files.get("layoutLWT"))
            print(f"update_temp_config_files: 文件路径已更新到temperature_config.json")
        else:
            print(f"update_temp_config_files: 跳过更新，条件不满足")
    
    def load_current_files_from_config(self):
        """從 temperature_config.json 配置檔載入上次選擇的檔案。

        驗證檔案是否仍然存在，若不存在則自動選擇第一個可用的檔案。
        """
        if self.current_folder_path:
            # 从temperature_config.json加载上次选择的文件
            if self.temp_config:
                saved_heat = self.temp_config.get_file_path("current_heat_file")
                saved_layout = self.temp_config.get_file_path("current_layout_file")
                saved_temp = self.temp_config.get_file_path("current_temp_file")
                saved_layout_xy = self.temp_config.get_file_path("current_layout_xy_file")
                saved_layout_lwt = self.temp_config.get_file_path("current_layout_lwt_file")
            else:
                # 如果temp_config未初始化，使用默认值
                saved_heat = saved_layout = saved_temp = None
                saved_layout_xy = saved_layout_lwt = None

            print(f"从配置文件加载的文件路径:")
            print(f"  current_heat_file: {saved_heat}")
            print(f"  current_layout_file: {saved_layout}")
            print(f"  current_temp_file: {saved_temp}")
            print(f"  current_layout_xy_file: {saved_layout_xy}")
            print(f"  current_layout_lwt_file: {saved_layout_lwt}")

            # 验证文件是否仍然存在，如果不存在则执行默认操作
            self._load_or_default_file("heat", saved_heat, "热力图")
            self._load_or_default_file("layout", saved_layout, "Layout图")
            self._load_or_default_file("heatTemp", saved_temp, "温度数据")
            self._load_or_default_file("layoutXY", saved_layout_xy, "元器件座標")
            self._load_or_default_file("layoutLWT", saved_layout_lwt, "元器件尺寸")
            
            print(f"文件选择完成: {self.current_files}")
    
    def _load_or_default_file(self, file_type, saved_file, display_name):
        """載入指定檔案類型，若配置的檔案不存在則使用預設的第一個可用檔案。

        參數：
            file_type (str): 檔案分類鍵值（"heat", "layout", "heatTemp", "layoutXY", "layoutLWT"）
            saved_file (str): 從配置中讀取的檔案名稱
            display_name (str): 用於日誌輸出的中文顯示名稱
        """
        if saved_file and saved_file in self.folder_files.get(file_type, []):
            # 配置的文件存在，使用配置的文件
            self.current_files[file_type] = saved_file
            print(f"✓ 恢复{display_name}选择: {saved_file}")
        elif self.folder_files.get(file_type):
            # 配置的文件不存在，使用默认操作：选择第一个可用的文件
            self.current_files[file_type] = self.folder_files[file_type][0]
            print(f"⚠ {display_name}配置不存在或文件已删除，自动选择第一个可用文件: {self.current_files[file_type]}")
            
            # 更新配置文件，保存默认选择的文件
            if self.temp_config:
                # 正确的配置键映射
                if file_type == "heat":
                    config_key = "current_heat_file"
                elif file_type == "layout":
                    config_key = "current_layout_file"
                elif file_type == "heatTemp":
                    config_key = "current_temp_file"
                elif file_type == "layoutXY":
                    config_key = "current_layout_xy_file"
                elif file_type == "layoutLWT":
                    config_key = "current_layout_lwt_file"
                else:
                    config_key = f"current_{file_type}_file"
                
                self.temp_config.set_file_path(config_key, self.current_files[file_type])
                print(f"已更新配置文件: {config_key} = {self.current_files[file_type]}")
        else:
            # 没有可用的文件
            self.current_files[file_type] = None
            print(f"⚠ 没有可用的{display_name}文件")
    
    def clear_old_data(self):
        """清空記憶體中的舊數據，在切換資料夾時呼叫。

        清除對齊點、座標變換器、標記框、圖片、溫度數據等，
        但不刪除磁碟上的任何檔案。
        """
        # 清空点位数据
        self.points_A = []
        self.points_B = []
        
        # 清空点转换器
        self.point_transformer = None
        
        # 清空标记矩形数据
        self.mark_rect_A = []
        self.mark_rect_B = []
        
        # 清空图片数据
        self.imageA = None
        self.imageB = None
        self.resized_imageA = None
        self.resized_imageB = None
        
        # 清空画布显示
        if hasattr(self, 'canvasA'):
            self.canvasA.delete("all")
        if hasattr(self, 'canvasB'):
            self.canvasB.delete("all")
        
        # 清空温度数据
        if hasattr(self, 'tempALoader'):
            self.tempALoader = None
        self.current_temp_file_path = None
        
        # 清空当前文件信息
        self.current_files = {"heat": None, "layout": None, "heatTemp": None, "layoutXY": None, "layoutLWT": None, "testReport": None}

        # 清空Layout数据
        self.layout_data = None
        self._category_warnings = set()

        print("已清空内存中的旧数据，准备加载新文件夹数据")
    
    
    def select_folder(self):
        """開啟資料夾選擇對話框，讓使用者選擇工作資料夾。

        選擇後會依序執行：儲存舊資料夾配置 -> 清空舊數據 -> 掃描新資料夾 ->
        更新檔案樹 -> 載入對齊點 -> 更新按鈕文字。
        """
        folder_path = filedialog.askdirectory(title="选择包含热力图和Layout图的文件夹")
        if folder_path:
            # 保存当前文件夹的文件选择
            self.save_current_files_to_config()
            
            # 清空旧的数据
            self.clear_old_data()
            
            self.current_folder_path = folder_path
            # 初始化温度配置管理器
            self.temp_config = TemperatureConfigManager(folder_path)
            self.save_folder_path()
            self.scan_folder_files()
            # 在扫描文件后更新温度配置管理器中的文件信息
            print(f"set_folder_path: 准备调用update_temp_config_files")
            self.update_temp_config_files()
            self.update_folder_display()
            self.update_folder_path_label()
            
            # 重新加载点位数据
            self.load_points()
            
            
            # 更新按钮文字
            folder_name = os.path.basename(folder_path)
            self.folder_control_button.config(text=f"隐藏文件夹Tab")
    
    def scan_folder_files(self):
        """掃描當前資料夾中的所有檔案，並根據檔案內容自動分類。

        委託 folder_scanner.scan_folder() 執行實際的分類邏輯，
        再將結果寫入 self.folder_files 和 self.current_files。
        分類完成後驗證 Layout 數據完整性並自動載入圖片。
        """
        if not self.current_folder_path:
            return

        try:
            # 委託 folder_scanner 執行分類
            self.folder_files, self._xlsx_columns_cache = scan_folder(self.current_folder_path)

            # 設定當前使用的檔案為最新的檔案
            for category in self.folder_files:
                if self.folder_files[category]:
                    self.current_files[category] = self.folder_files[category][0]

            # 驗證 Layout 數據檔案完整性，並記錄有問題的分類
            self._category_warnings = set()
            warning_msg, warned_categories = validate_layout_data(self.folder_files, self._xlsx_columns_cache)
            if warning_msg:
                self._category_warnings.update(warned_categories)
                self.root.after(200, lambda: messagebox.showwarning("Layout 數據檢查", warning_msg))

            # 掃描完成後，自動載入可用的圖片
            self.auto_load_images()

            # 根據是否偵測到測試報告來啟用/停用按鈕
            if hasattr(self, 'test_report_button'):
                if self.folder_files.get("testReport"):
                    self.test_report_button.config(state=tk.NORMAL)
                else:
                    self.test_report_button.config(state=tk.DISABLED)

            # 添加調試資訊
            print(f"scan_folder_files: 扫描完成，current_files: {self.current_files}")
        except Exception as e:
            print(f"扫描文件夹时出错: {e}")

    def auto_load_images(self):
        """自動載入各分類中當前選用的檔案。

        執行順序：
            1. 從配置恢復上次選擇的檔案
            2. 同步載入熱力圖和 Layout 圖（快速顯示）
            3. 同步載入對齊點數據
            4. 非同步載入溫度數據（子執行緒，避免阻塞 UI）
            5. 非同步載入 Layout 數據（子執行緒）
        """
        try:
            # 首先尝试从配置恢复上次选择的文件
            self.load_current_files_from_config()
            
            # 优先加载图片（同步加载，快速显示）
            if self.current_files["heat"]:
                heat_image_path = os.path.join(self.current_folder_path, self.current_files["heat"])
                self.set_image(heat_image_path, 0)
                print(f"自动加载热力图: {self.current_files['heat']}")
            
            if self.current_files["layout"]:
                layout_image_path = os.path.join(self.current_folder_path, self.current_files["layout"])
                self.set_image(layout_image_path, 1)
                print(f"自动加载Layout图: {self.current_files['layout']}")
            
            # 加载点位数据（同步加载，快速显示）
            self.load_points()
            
            # 温度数据异步加载（在子线程中）
            if self.current_files["heatTemp"]:
                temp_file_path = os.path.join(self.current_folder_path, self.current_files["heatTemp"])
                self.load_temperature_file_async(temp_file_path)
            
            # Layout数据异步加载（在子线程中）
            if self.folder_files.get("layoutXY") and self.folder_files.get("layoutLWT"):
                # 自动加载所有layout数据文件
                self.load_all_layout_data_async()
                
        except Exception as e:
            print(f"自动加载图片时出错: {e}")
    
    def update_folder_display(self):
        """更新左側面板的資料夾檔案樹狀列表顯示。

        清空現有內容後，依分類（熱力圖/Layout圖/溫度數據/Layout數據）
        重新建構樹狀結構。當前選用的檔案會以粗體高亮顯示。
        """
        if hasattr(self, 'folder_tree'):
            # 记录当前展开状态
            expanded_items = []
            for item in self.folder_tree.get_children():
                if self.folder_tree.item(item, "open"):
                    expanded_items.append(self.folder_tree.item(item, "text"))
            
            # 清空现有内容
            for item in self.folder_tree.get_children():
                self.folder_tree.delete(item)
            
            # 固定六個分類，永遠全部顯示
            all_categories = ["heat", "layout", "heatTemp", "layoutXY", "layoutLWT", "testReport"]
            category_names = {"heat": "熱力圖", "layout": "Layout圖", "heatTemp": "溫度數據", "layoutXY": "元器件座標", "layoutLWT": "元器件尺寸", "testReport": "測試報告"}
            category_spaces = {"heat": 32, "layout": 32, "heatTemp": 29, "layoutXY": 27, "layoutLWT": 27, "testReport": 27}

            for category in all_categories:
                files = self.folder_files.get(category, [])
                category_name = category_names[category]
                # 沒有檔案 → ❗；有驗證警告 → ❗；否則 → ✅
                warned = hasattr(self, '_category_warnings') and category in self._category_warnings
                has_warning = (not files) or warned
                status_icon = " ❗" if has_warning else " ✅"
                base_text = f"{category_name} ({len(files)}){status_icon}"
                display_text = f"{base_text:<{category_spaces[category]}}📁"
                category_item = self.folder_tree.insert("", "end", text=display_text, values=(category, ""))

                # 自动展开所有分类
                self.folder_tree.item(category_item, open=True)

                if files:
                    for filename in files:
                        if filename == self.current_files.get(category):
                            item = self.folder_tree.insert(category_item, "end", text=filename, values=(category, filename))
                            self.folder_tree.item(item, tags=("bold",))
                        else:
                            self.folder_tree.insert(category_item, "end", text=filename, values=(category, filename))
                else:
                    # 無檔案時顯示「無」
                    self.folder_tree.insert(category_item, "end", text="無", values=(category, ""))
    
    def _on_tree_motion(self, event):
        """處理檔案樹狀列表的滑鼠移動事件，顯示 layoutXY / layoutLWT 的欄位說明 tooltip。"""
        item = self.folder_tree.identify_row(event.y)
        if not item:
            self._hide_tree_tooltip()
            return

        # 取得該項目的 category
        values = self.folder_tree.item(item, "values")
        if not values or not values[0]:
            self._hide_tree_tooltip()
            return

        category = values[0]
        filename = values[1] if len(values) > 1 else ""

        # 定義需要顯示 tooltip 的類別及內容
        tooltip_texts = {
            "heat": (
                "熱力圖影像檔案\n"
                "支援格式：.jpg, .jpeg, .png\n"
                "辨識條件：HSV 飽和度 > 80 且色調變異數 > 1000"
            ),
            "layout": (
                "Layout 佈局圖影像檔案\n"
                "支援格式：.jpg, .jpeg, .png\n"
                "辨識條件：黑色像素比例 > 60%"
            ),
            "heatTemp": (
                "熱力圖溫度數據\n"
                "支援格式：.csv（Tab/逗號分隔）或 .xlsx\n"
                "辨識條件：列數≥50 的寬數值矩陣，數字佔比>90%\n"
                "可容忍前幾行文字表頭（自動跳過）"
            ),
            "layoutXY": (
                "元器件座標與角度檔案\n"
                "支援格式：.xlsx\n"
                "歸類條件：同時含有 X, Y 欄位\n"
                "完整欄位：RefDes, Orient., X, Y\n"
                "缺少 RefDes 或 Orient. 時會顯示警告"
            ),
            "layoutLWT": (
                "元器件描述與長寬高檔案\n"
                "支援格式：.xlsx\n"
                "歸類條件：同時含有 L, W, T 欄位\n"
                "完整欄位：RefDes, 对象描述, L, W, T\n"
                "缺少 RefDes 或 对象描述 時會顯示警告"
            ),
            "testReport": (
                "測試報告檔案\n"
                "支援格式：.xlsx\n"
                "辨識條件：轉大寫含有 HIGH 字眼的 Sheet"
            ),
        }

        if filename:
            # 檔案項目：tooltip 顯示完整檔案名稱
            tooltip_content = filename
        else:
            # 類別標題：tooltip 顯示類別說明
            tooltip_content = tooltip_texts.get(category)
            if not tooltip_content:
                self._hide_tree_tooltip()
                return

        # 如果已經是同一個 item 的 tooltip，只更新位置
        if self._tree_tooltip and self._tree_tooltip_item == item:
            x = event.x_root + 15
            y = event.y_root + 10
            self._tree_tooltip.wm_geometry(f"+{x}+{y}")
            return

        # 隱藏舊的 tooltip，建立新的
        self._hide_tree_tooltip()
        self._tree_tooltip_item = item

        tw = tk.Toplevel(self.folder_tree)
        tw.wm_overrideredirect(True)
        label = tk.Label(
            tw, text=tooltip_content,
            justify=tk.LEFT, background="#FFFFCC", foreground="#000000",
            relief=tk.SOLID, borderwidth=1, font=("Arial", 9),
            padx=8, pady=6
        )
        label.pack()
        x = event.x_root + 15
        y = event.y_root + 10
        tw.wm_geometry(f"+{x}+{y}")
        tw.lift()
        self._tree_tooltip = tw

    def _hide_tree_tooltip(self, event=None):
        """隱藏檔案樹狀列表的 tooltip。"""
        if self._tree_tooltip:
            self._tree_tooltip.destroy()
            self._tree_tooltip = None
            self._tree_tooltip_item = None

    def update_folder_path_label(self):
        """更新左側面板頂部的「當前資料夾」路徑標籤文字。"""
        if hasattr(self, 'folder_path_label'):
            if self.current_folder_path:
                # 显示文件夹名称而不是完整路径
                folder_name = os.path.basename(self.current_folder_path)
                self.folder_path_label.config(text=f"当前文件夹：{folder_name}")
            else:
                self.folder_path_label.config(text="当前文件夹：未选择")
    def on_file_click(self, event):
        """處理檔案樹狀列表的單擊事件。

        根據點擊位置判斷行為：
            - 點擊分類標題右側的資料夾圖示：開啟檔案選擇對話框
            - 點擊分類標題文字區域：折疊/展開該分類
            - 點擊檔案項目：切換到該檔案（載入並更新顯示）

        參數：
            event: Tkinter 滑鼠事件物件（包含 x, y 座標等資訊）
        """
        # 通过点击位置确定实际点击的项目，而不是使用selection()
        item = self.folder_tree.identify_row(event.y)
        if not item:
            return
            
        values = self.folder_tree.item(item, "values")
        item_text = self.folder_tree.item(item, "text")
        
        # 检查是否是父标题项（分类项）
        if len(values) >= 2 and values[0] and not values[1]:  # 父标题项：有category但没有filename
            category = values[0]
            
            # 检查是否点击了选择图标
            if "📁" in item_text:
                # 计算点击位置，判断是否点击在选择图标区域
                bbox = self.folder_tree.bbox(item)
                if bbox:
                    x, y, width, height = bbox
                    # 计算图标位置（在文本的右侧）
                    # 由于使用了固定宽度，图标应该在右侧
                    icon_start_x = x + width - 25  # 图标大约占25像素宽度
                    print(f"点击检测: event.x={event.x}, bbox=({x},{y},{width},{height}), icon_start_x={icon_start_x}")
                    if event.x >= icon_start_x:  # 点击在图标区域
                        print(f"点击了图标，开始选择文件: {category}")
                        # 立即执行文件选择，阻止默认行为
                        self.select_and_replace_current_file(category)
                        return "break"  # 阻止事件继续传播
                    else:
                        print(f"点击了父标题文本区域，执行折叠/展开操作")
                        # 普通点击父标题，执行折叠/展开操作
                        return
        
        # 处理文件项
        elif len(values) >= 2 and values[0] and values[1]:  # 文件项：有category和filename
            category = values[0]
            filename = values[1]
            
            # layoutXY/layoutLWT 類型的文件項不可切換，不觸發任何動作
            if category in ("layoutXY", "layoutLWT"):
                print(f"{category}文件项不可切换: {filename}")
                return "break"  # 阻止事件继续传播
            
            # 处理文件切换（单击文件名区域）
            file_path = os.path.join(self.current_folder_path, filename)
            
            # 更新当前使用的文件
            self.current_files[category] = filename
            
            # 更新温度配置管理器
            self.update_temp_config_files()
            
            # 保存当前选择的文件到配置
            self.save_current_files_to_config()
            
            # 加载新文件
            if category == "heat":
                self.set_image(file_path, 0)
                print(f"切换到热力图: {filename}")
                # 切换热力图后，清空对齐点数据并重新加载
                self.clear_and_reload_points()
            elif category == "layout":
                self.set_image(file_path, 1)
                print(f"切换到Layout图: {filename}")
                # 切换Layout图后，清空对齐点数据并重新加载
                self.clear_and_reload_points()
            elif category == "heatTemp":
                self.load_temperature_file(file_path)
                print(f"切换到温度数据: {filename}")
            elif category in ("layoutXY", "layoutLWT"):
                self.load_all_layout_data_async()
                print(f"切换到{category}数据: {filename}")
            
            # 刷新文件夹显示，更新加粗标记
            self.update_folder_display()
    
    def select_and_replace_current_file(self, category):
        """開啟檔案選擇對話框，選擇新檔案並替換當前使用的資源。

        若選擇的檔案不在當前資料夾中，會自動複製過來。不會刪除原檔案。

        參數：
            category (str): 檔案分類（"heat", "layout", "heatTemp", "layoutXY", "layoutLWT"）
        """
        try:
            print(f"select_and_replace_current_file 被调用，category = {category}")
            # 根据分类设置文件类型过滤器
            if category == "heat":
                filetypes = [("图片文件", "*.jpg *.jpeg *.png"), ("所有文件", "*.*")]
                title = "选择热力图文件"
                print(f"设置热力图文件过滤器: {filetypes}")
            elif category == "layout":
                filetypes = [("图片文件", "*.jpg *.jpeg *.png"), ("所有文件", "*.*")]
                title = "选择Layout图文件"
                print(f"设置Layout图文件过滤器: {filetypes}")
            elif category == "heatTemp":
                filetypes = [("数据文件", "*.csv *.xlsx"), ("所有文件", "*.*")]
                title = "选择温度数据文件"
                print(f"设置温度数据文件过滤器: {filetypes}")
            elif category == "layoutXY":
                filetypes = [("Excel文件", "*.xlsx")]
                title = "選擇元器件座標檔案 (需含 RefDes, Orient., X, Y)"
                print(f"设置元器件座標文件过滤器: {filetypes}")
            elif category == "layoutLWT":
                filetypes = [("Excel文件", "*.xlsx")]
                title = "選擇元器件尺寸檔案 (需含 RefDes, L, W, T, 对象描述)"
                print(f"设置元器件尺寸文件过滤器: {filetypes}")
            else:
                print(f"未知分类: {category}")
                return
            
            # 打开文件选择对话框
            print(f"准备打开文件对话框: title={title}, filetypes={filetypes}")
            file_path = filedialog.askopenfilename(
                title=title,
                filetypes=filetypes,
                initialdir=self.current_folder_path
            )
            print(f"文件对话框返回: {file_path}")
            
            if file_path:
                # 获取新文件名
                new_filename = os.path.basename(file_path)
                new_file_path = os.path.join(self.current_folder_path, new_filename)
                
                # 复制新文件到当前文件夹（如果文件不存在）
                if not os.path.exists(new_file_path):
                    import shutil
                    shutil.copy2(file_path, new_file_path)
                    print(f"已复制新文件: {new_filename} 到当前文件夹")
                else:
                    print(f"文件已存在: {new_filename}")
                
                # 更新当前使用的文件
                self.current_files[category] = new_filename
                
                # 保存当前选择的文件到配置
                self.save_current_files_to_config()
                
                # 重新扫描文件夹文件
                self.scan_folder_files()
                
                # 刷新文件夹显示
                self.update_folder_display()
                
                # 加载新文件到内存（替换当前使用的资源）
                if category == "heat":
                    self.set_image(new_file_path, 0)
                    print(f"已加载热力图: {new_filename}")
                    # 替换热力图后，清空对齐点数据并重新加载
                    self.clear_and_reload_points()
                elif category == "layout":
                    self.set_image(new_file_path, 1)
                    print(f"已加载Layout图: {new_filename}")
                    # 替换Layout图后，清空对齐点数据并重新加载
                    self.clear_and_reload_points()
                elif category == "heatTemp":
                    self.load_temperature_file(new_file_path)
                    print(f"已加载温度数据: {new_filename}")
                elif category in ("layoutXY", "layoutLWT"):
                    self.load_all_layout_data_async()
                    print(f"已加载{category}数据: {new_filename}")
                
                # 显示成功消息
                show_toast(
                    title='文件替换成功',
                    message=f'已切换到{new_filename}',
                    duration=3000,
                    toast_type='success'
                )
                
        except Exception as e:
            print(f"替换文件时出错: {e}")
            from tkinter import messagebox
            messagebox.showerror("错误", f"替换文件失败: {e}")
    
    def select_and_replace_file(self, category, old_filename):
        """選擇新檔案並替換（刪除）舊檔案。

        與 select_and_replace_current_file 不同，本方法會刪除舊檔案。

        參數：
            category (str): 檔案分類（"heat", "pcb", "heatTemp"）
            old_filename (str): 要被替換的舊檔案名稱
        """
        try:
            # 根据分类设置文件类型过滤器
            if category == "heat":
                filetypes = [("图片文件", "*.jpg *.jpeg *.png"), ("所有文件", "*.*")]
                title = "选择热力图文件"
            elif category == "pcb":
                filetypes = [("图片文件", "*.jpg *.jpeg *.png"), ("所有文件", "*.*")]
                title = "选择Layout图文件"
            elif category == "heatTemp":
                filetypes = [("数据文件", "*.csv *.xlsx"), ("所有文件", "*.*")]
                title = "选择温度数据文件"
            else:
                return
            
            # 打开文件选择对话框
            file_path = filedialog.askopenfilename(
                title=title,
                filetypes=filetypes
            )
            
            if file_path:
                # 获取新文件名
                new_filename = os.path.basename(file_path)
                old_file_path = os.path.join(self.current_folder_path, old_filename)
                new_file_path = os.path.join(self.current_folder_path, new_filename)
                
                # 如果新文件名与旧文件名不同，需要替换
                if new_filename != old_filename:
                    # 删除旧文件
                    if os.path.exists(old_file_path):
                        os.remove(old_file_path)
                        print(f"已删除旧文件: {old_filename}")
                    
                    # 复制新文件到当前文件夹
                    import shutil
                    shutil.copy2(file_path, new_file_path)
                    print(f"已复制新文件: {new_filename} 到当前文件夹")
                else:
                    # 文件名相同，直接覆盖
                    import shutil
                    shutil.copy2(file_path, new_file_path)
                    print(f"已覆盖文件: {new_filename}")
                
                # 更新当前使用的文件
                self.current_files[category] = new_filename
                
                # 保存当前选择的文件到配置
                self.save_current_files_to_config()
                
                # 重新扫描文件夹文件
                self.scan_folder_files()
                
                # 刷新文件夹显示
                self.update_folder_display()
                
                # 加载新文件
                if category == "heat":
                    self.set_image(new_file_path, 0)
                    print(f"已加载热力图: {new_filename}")
                elif category == "pcb":
                    self.set_image(new_file_path, 1)
                    print(f"已加载Layout图: {new_filename}")
                elif category == "heatTemp":
                    self.load_temperature_file(new_file_path)
                    print(f"已加载温度数据: {new_filename}")
                
                # 显示成功消息
                show_toast(
                    title='文件替换成功',
                    message=f'已替换{old_filename}为{new_filename}',
                    duration=3000,
                    toast_type='success'
                )
                
        except Exception as e:
            print(f"替换文件时出错: {e}")
            from tkinter import messagebox
            messagebox.showerror("错误", f"替换文件失败: {e}")
    
    def select_and_copy_file(self, category):
        """選擇外部檔案並複製到當前資料夾中，然後設為當前使用的檔案。

        參數：
            category (str): 檔案分類（"heat", "pcb", "heatTemp"）
        """
        try:
            # 根据分类设置文件类型过滤器
            if category == "heat":
                filetypes = [("图片文件", "*.jpg *.jpeg *.png"), ("所有文件", "*.*")]
                title = "选择热力图文件"
            elif category == "pcb":
                filetypes = [("图片文件", "*.jpg *.jpeg *.png"), ("所有文件", "*.*")]
                title = "选择Layout图文件"
            elif category == "heatTemp":
                filetypes = [("数据文件", "*.csv *.xlsx"), ("所有文件", "*.*")]
                title = "选择温度数据文件"
            else:
                return
            
            # 打开文件选择对话框
            file_path = filedialog.askopenfilename(
                title=title,
                filetypes=filetypes
            )
            
            if file_path:
                # 获取文件名
                filename = os.path.basename(file_path)
                target_path = os.path.join(self.current_folder_path, filename)
                
                # 复制文件到当前文件夹
                import shutil
                shutil.copy2(file_path, target_path)
                print(f"已复制文件: {filename} 到当前文件夹")
                
                # 更新当前使用的文件
                self.current_files[category] = filename
                
                # 更新温度配置管理器
                self.update_temp_config_files()
                
                # 保存当前选择的文件到配置
                self.save_current_files_to_config()
                
                # 重新扫描文件夹文件
                self.scan_folder_files()
                
                # 刷新文件夹显示
                self.update_folder_display()
                
                # 加载新复制的文件
                if category == "heat":
                    self.set_image(target_path, 0)
                    print(f"已加载热力图: {filename}")
                elif category == "pcb":
                    self.set_image(target_path, 1)
                    print(f"已加载Layout图: {filename}")
                elif category == "heatTemp":
                    self.load_temperature_file(target_path)
                    print(f"已加载温度数据: {filename}")
                
                # 显示成功消息
                show_toast(
                    title='文件选择成功',
                    message=f'已选择并复制{filename}到当前文件夹',
                    duration=3000,
                    toast_type='success'
                )
                
        except Exception as e:
            print(f"选择文件时出错: {e}")
            from tkinter import messagebox
            messagebox.showerror("错误", f"选择文件失败: {e}")
    
    def _validate_image_temp_dimensions(self):
        """比對熱力圖影像與溫度數據的解析度尺寸是否一致。

        當兩者都已載入時，比較熱力圖的像素尺寸與溫度矩陣的行列數。
        若不一致則彈出警告對話框，提示使用者溫度座標對應可能不正確。
        """
        if self.imageA is None:
            return
        if not hasattr(self, 'tempALoader') or self.tempALoader is None:
            return
        temp_data = self.tempALoader.get_tempA()
        if temp_data is None:
            return

        img_w, img_h = self.imageA.size        # PIL: (width, height)
        temp_h, temp_w = temp_data.shape        # NumPy: (rows, cols) = (height, width)

        if img_w != temp_w or img_h != temp_h:
            self._category_warnings.add("heat")
            self._category_warnings.add("heatTemp")
            self.update_folder_display()
            msg = (
                f"熱力圖影像與溫度數據的解析度不一致：\n\n"
                f"  熱力圖影像：{img_w} × {img_h}\n"
                f"  溫度數據：    {temp_w} × {temp_h}\n\n"
                f"溫度座標對應可能會不正確，請確認檔案是否匹配。"
            )
            messagebox.showwarning("解析度不一致", msg)
        else:
            # 解析度一致時，移除之前可能存在的警告
            self._category_warnings.discard("heat")
            self._category_warnings.discard("heatTemp")
            self.update_folder_display()

    def load_temperature_file(self, file_path):
        """同步載入溫度數據檔案（CSV 或 XLSX 格式）。

        建立 TempLoader 實例並設定全域溫度檔案路徑。
        載入後會自動比對熱力圖影像的解析度是否一致。

        參數：
            file_path (str): 溫度數據檔案的完整路徑
        """
        try:
            if file_path.lower().endswith(('.csv', '.xlsx')):
                # 设置全局温度文件路径
                self.current_temp_file_path = file_path
                # 更新TempLoader的文件路径
                self.tempALoader = TempLoader(file_path)
                print(f"已加载温度数据文件: {file_path}")
                self._validate_image_temp_dimensions()
            else:
                print(f"不支持的文件格式: {file_path}")
        except Exception as e:
            messagebox.showerror("错误", f"加载温度数据文件失败: {e}")

    def load_temperature_file_async(self, file_path):
        """在子執行緒中非同步載入溫度數據檔案，避免阻塞 UI。

        參數：
            file_path (str): 溫度數據檔案的完整路徑
        """
        def load_temp_data():
            try:
                print(f"开始异步加载温度数据: {file_path}")
                if file_path.lower().endswith(('.csv', '.xlsx')):
                    # 设置全局温度文件路径
                    self.current_temp_file_path = file_path
                    # 更新TempLoader的文件路径
                    self.tempALoader = TempLoader(file_path)
                    print(f"异步加载温度数据完成: {file_path}")
                    # 在主執行緒中比對解析度
                    self.root.after(0, self._validate_image_temp_dimensions)
                else:
                    print(f"不支持的文件格式: {file_path}")
            except Exception as e:
                print(f"异步加载温度数据失败: {e}")
                # 在主线程中显示错误信息
                self.root.after(0, lambda: messagebox.showerror("错误", f"加载温度数据文件失败: {e}"))

            # 初始化yolo
            # if not hasattr(self, 'yolo'):
            #     self.yolo = YOLOv8Instance()
        
        # 在子线程中加载温度数据
        temp_thread = threading.Thread(target=load_temp_data, daemon=True)
        temp_thread.start()
    
    def load_layout_data_async(self, file_path):
        """在子執行緒中非同步載入單一 Layout 數據檔案。

        參數：
            file_path (str): Layout 數據檔案的完整路徑
        """
        def load_layout_data():
            try:
                print(f"开始异步加载Layout数据: {file_path}")
                self.layout_data = self.parse_layout_data(file_path)
                print(f"异步加载Layout数据完成: {file_path}")
            except Exception as e:
                print(f"异步加载Layout数据失败: {e}")
                # 在主线程中显示错误信息
                self.root.after(0, lambda: messagebox.showerror("错误", f"加载Layout数据文件失败: {e}"))
        
        # 在子线程中加载Layout数据
        layout_thread = threading.Thread(target=load_layout_data, daemon=True)
        layout_thread.start()
    
    def load_all_layout_data_async(self):
        """在子執行緒中非同步載入資料夾內的元器件座標檔和尺寸檔。

        直接使用已分類的 folder_files["layoutXY"] 和 folder_files["layoutLWT"]，
        解析後合併為完整的元器件資訊列表。
        """
        def load_all_layout_data():
            try:
                print(f"开始异步加载所有Layout数据文件...")
                xy_files = self.folder_files.get("layoutXY", [])
                lwt_files = self.folder_files.get("layoutLWT", [])
                if not xy_files or not lwt_files:
                    print("Layout數據不完整：缺少座標檔或尺寸檔")
                    return

                c_file = os.path.join(self.current_folder_path, xy_files[0])
                c_item_file = os.path.join(self.current_folder_path, lwt_files[0])

                print(f"元器件座標檔: {xy_files[0]}, 元器件尺寸檔: {lwt_files[0]}")

                # 解析文件并计算C_info
                self.layout_data = self.parse_all_layout_data(c_file, c_item_file)
                print(f"异步加载所有Layout数据完成，共{len(self.layout_data) if self.layout_data else 0}个元器件")

            except Exception as e:
                print(f"异步加载所有Layout数据失败: {e}")
                # 在主线程中显示错误信息
                self.root.after(0, lambda: messagebox.showerror("错误", f"加载Layout数据文件失败: {e}"))

        # 在子线程中加载Layout数据
        layout_thread = threading.Thread(target=load_all_layout_data, daemon=True)
        layout_thread.start()
    
    def parse_all_layout_data(self, c_file, c_item_file):
        """解析元器件座標檔和尺寸檔，回傳合併後的元器件資訊列表。

        根據 RefDes 欄位合併兩個檔案的數據，計算每個元器件的邊界框。

        參數：
            c_file (str): 元器件座標檔路徑（含 RefDes, Orient., X, Y）
            c_item_file (str): 元器件尺寸檔路徑（含 RefDes, L, W, T, 对象描述）

        回傳：
            list 或 None: 元器件資訊列表，每項為 dict 包含
                {RefDes, left, top, right, bottom, X, Y, L, W, T, Orient.}
        """
        try:
            if not c_file or not c_item_file:
                print(f"未找到合适的Layout数据文件")
                return None

            # 读取C.xlsx文件
            c_df = pd.read_excel(c_file)
            print(f"C文件字段: {c_df.columns.tolist()}")
            
            # 读取C_item.xlsx文件
            c_item_df = pd.read_excel(c_item_file)
            print(f"C_item文件字段: {c_item_df.columns.tolist()}")
            
            # 检查必需字段
            required_c_fields = ['RefDes', 'Orient.', 'X', 'Y']
            required_item_fields = ['RefDes', 'L', 'W', 'T', '对象描述']
            
            for field in required_c_fields:
                if field not in c_df.columns:
                    print(f"C文件缺少必需字段: {field}")
                    return None
            
            for field in required_item_fields:
                if field not in c_item_df.columns:
                    print(f"C_item文件缺少必需字段: {field}")
                    return None
            
            # 解析数据
            c_info = []
            
            for _, row in c_df.iterrows():
                refdes = row['RefDes']
                x = row['X']
                y = row['Y']
                orient = row['Orient.']
                
                # 在C_item文件中查找对应的尺寸信息
                item_match = c_item_df[c_item_df['RefDes'] == refdes]
                if not item_match.empty:
                    item_row = item_match.iloc[0]
                    l = item_row['L']
                    w = item_row['W']
                    t = item_row['T']

                    # 提取「对象描述」欄位（如果存在）
                    description = item_row.get('对象描述', '')
                    if pd.isna(description):
                        description = ''

                    # 计算边界框（考虑旋转角度）
                    if orient == 0 or pd.isna(orient):
                        # 如果角度为0或NaN，使用简单计算
                        left = x - l/2
                        top = y - w/2
                        right = x + l/2
                        bottom = y + w/2
                    else:
                        # 使用旋转计算
                        left, top, right, bottom = self.calculate_rotated_bounding_box(x, y, l, w, orient)

                    c_info.append({
                        'RefDes': refdes,
                        'left': left,
                        'top': top,
                        'right': right,
                        'bottom': bottom,
                        'X': x,
                        'Y': y,
                        'L': l,
                        'W': w,
                        'T': t,
                        'Orient.': orient,
                        'Description': description
                    })
                # else:
                    # print(f"未找到RefDes {refdes} 对应的尺寸信息，跳过")
            
            print(f"成功解析 {len(c_info)} 个元器件信息")
            return c_info
            
        except Exception as e:
            print(f"解析Layout数据失败: {e}")
            return None
    
    def calculate_rotated_bounding_box(self, x, y, length, width, angle_deg):
        """計算元器件旋轉後的軸對齊邊界框 (AABB)。

        根據元器件的中心座標、長寬尺寸和旋轉角度，計算旋轉後
        能完整包含元器件的最小軸對齊矩形。

        參數：
            x (float): 元器件中心 X 座標 (mm)
            y (float): 元器件中心 Y 座標 (mm)
            length (float): 元器件長度 (mm)
            width (float): 元器件寬度 (mm)
            angle_deg (float): 旋轉角度（度），正值為順時針

        回傳：
            tuple: (left, top, right, bottom) 旋轉後的邊界框座標
        """
        import math
        
        # 将角度转换为弧度
        angle_rad = math.radians(angle_deg)
        
        # 计算半长和半宽
        half_length = length / 2
        half_width = width / 2
        
        # 计算四个角点相对于中心的坐标
        corners = [
            (-half_length, -half_width),  # 左下
            (half_length, -half_width),   # 右下
            (half_length, half_width),    # 右上
            (-half_length, half_width)    # 左上
        ]
        
        # 旋转每个角点
        rotated_corners = []
        for corner_x, corner_y in corners:
            # 旋转矩阵计算
            rotated_x = corner_x * math.cos(angle_rad) - corner_y * math.sin(angle_rad)
            rotated_y = corner_x * math.sin(angle_rad) + corner_y * math.cos(angle_rad)
            rotated_corners.append((rotated_x, rotated_y))
        
        # 计算旋转后的边界框
        x_coords = [x + corner[0] for corner in rotated_corners]
        y_coords = [y + corner[1] for corner in rotated_corners]
        
        left = min(x_coords)
        right = max(x_coords)
        top = min(y_coords)
        bottom = max(y_coords)
        
        return left, top, right, bottom
    
    def parse_layout_data(self, file_path):
        """解析 Layout 數據檔案，回傳元器件資訊列表 (C_info)。

        先從已識別的 layoutXY 和 layoutLWT 檔案中取得 C 檔和 C_item 檔，
        若未找到則回退到資料夾掃描模式。

        參數：
            file_path (str): Layout 數據檔案路徑（用於取得資料夾位置）

        回傳：
            list 或 None: 元器件資訊列表
        """
        try:
            # 直接使用已经识别出的layout数据文件
            folder_path = os.path.dirname(file_path)
            c_file = None
            c_item_file = None
            
            # 從已分類的 layoutXY / layoutLWT 取得檔案
            if hasattr(self, 'folder_files'):
                xy_files = self.folder_files.get('layoutXY', [])
                lwt_files = self.folder_files.get('layoutLWT', [])
                if xy_files:
                    c_file = os.path.join(folder_path, xy_files[0])
                    print(f"使用已分類的元器件座標檔: {xy_files[0]}")
                if lwt_files:
                    c_item_file = os.path.join(folder_path, lwt_files[0])
                    print(f"使用已分類的元器件尺寸檔: {lwt_files[0]}")
            
            # 如果还是没有找到，回退到原来的查找方式
            if not c_file or not c_item_file:
                print(f"从已识别文件中未找到合适的文件，回退到文件夹扫描...")
                for filename in os.listdir(folder_path):
                    if filename.lower() == 'c.xlsx':
                        c_file = os.path.join(folder_path, filename)
                    elif filename.lower() == 'c_item.xlsx':
                        c_item_file = os.path.join(folder_path, filename)
            
            if not c_file or not c_item_file:
                print(f"未找到合适的Layout数据文件")
                return None
            
            # 读取C.xlsx文件
            c_df = pd.read_excel(c_file)
            print(f"C.xlsx字段: {c_df.columns.tolist()}")
            
            # 读取C_item.xlsx文件
            c_item_df = pd.read_excel(c_item_file)
            print(f"C_item.xlsx字段: {c_item_df.columns.tolist()}")
            
            # 检查必需字段
            required_c_fields = ['RefDes', 'Orient.', 'X', 'Y']
            required_item_fields = ['RefDes', 'L', 'W', 'T', '对象描述']
            
            for field in required_c_fields:
                if field not in c_df.columns:
                    print(f"C.xlsx缺少必需字段: {field}")
                    return None
            
            for field in required_item_fields:
                if field not in c_item_df.columns:
                    print(f"C_item.xlsx缺少必需字段: {field}")
                    return None
            
            # 解析数据
            c_info = []
            
            for _, row in c_df.iterrows():
                refdes = row['RefDes']
                x = row['X']
                y = row['Y']
                orient = row['Orient.']
                
                # 在C_item中查找对应的RefDes
                item_match = c_item_df[c_item_df['RefDes'] == refdes]
                
                if len(item_match) > 0:
                    item_row = item_match.iloc[0]
                    l = item_row['L']  # 长
                    w = item_row['W']  # 宽
                    t = item_row['T']  # 高

                    # 提取「对象描述」欄位（如果存在）
                    description = item_row.get('对象描述', '')
                    if pd.isna(description):
                        description = ''

                    # 计算外接矩形的四个角点坐标（考虑旋转）
                    left, top, right, bottom = self.calculate_rotated_rectangle(x, y, l, w, orient)

                    c_info.append({
                        'RefDes': refdes,
                        'left': left,
                        'top': top,
                        'right': right,
                        'bottom': bottom,
                        'X': x,
                        'Y': y,
                        'L': l,
                        'W': w,
                        'T': t,
                        'Orient.': orient,
                        'Description': description
                    })
                # else:
                    # print(f"未找到RefDes {refdes} 对应的尺寸信息，跳过")
            
            print(f"成功解析 {len(c_info)} 个元器件信息")
            return c_info
            
        except Exception as e:
            print(f"解析Layout数据失败: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def calculate_rotated_rectangle(self, x, y, l, t, orient):
        """計算元器件旋轉後的外接矩形邊界座標。

        將矩形的四個角點進行旋轉變換，再取外接矩形的邊界。

        參數：
            x (float): 中心 X 座標
            y (float): 中心 Y 座標
            l (float): 長度
            t (float): 高度
            orient (float): 旋轉角度（度）

        回傳：
            tuple: (left, top, right, bottom) 外接矩形邊界
        """
        try:
            # 将旋转角度转换为弧度
            # orient为.270表示顺时针旋转270度，orient为-270表示逆时针旋转270度
            angle_rad = math.radians(float(orient))
            
            # 计算矩形的半长和半高
            half_l = l / 2
            half_t = t / 2
            
            # 原始矩形的四个角点（相对于中心点）
            corners = [
                (-half_l, -half_t),  # 左下
                (half_l, -half_t),   # 右下
                (half_l, half_t),    # 右上
                (-half_l, half_t)    # 左上
            ]
            
            # 应用旋转变换
            rotated_corners = []
            for corner_x, corner_y in corners:
                # 旋转公式
                new_x = corner_x * math.cos(angle_rad) - corner_y * math.sin(angle_rad)
                new_y = corner_x * math.sin(angle_rad) + corner_y * math.cos(angle_rad)
                rotated_corners.append((new_x + x, new_y + y))
            
            # 计算外接矩形的边界
            x_coords = [corner[0] for corner in rotated_corners]
            y_coords = [corner[1] for corner in rotated_corners]
            
            left = min(x_coords)
            top = min(y_coords)
            right = max(x_coords)
            bottom = max(y_coords)
            
            return left, top, right, bottom
            
        except Exception as e:
            print(f"计算旋转矩形失败: {e}")
            # 如果计算失败，返回未旋转的矩形
            return x - l/2, y - t/2, x + l/2, y + t/2

    def update_magnifier_point(self):
        """更新放大鏡元件中的對齊點資料，使放大鏡能正確顯示點位標記。"""
        if hasattr(self, 'canvasA_magnifier') and self.canvasA_magnifier:
            self.canvasA_magnifier.update_points(self.points_A)
        if hasattr(self, 'canvasB_magnifier') and self.canvasB_magnifier:
            self.canvasB_magnifier.update_points(self.points_B)
    def init_magnifier(self):
        """初始化左右畫布的放大鏡元件。

        先清除舊的放大鏡實例，再根據當前縮放後的圖片建立新的放大鏡。
        僅在圖片已載入的情況下才會建立。
        """
        self.clean_magnifier()
        # 检查图片是否已加载
        if hasattr(self, 'resized_imageA') and hasattr(self, 'resized_imageB') and self.resized_imageA and self.resized_imageB:
            self.canvasA_magnifier = ImageMagnifier(self.canvasA, self.resized_imageA, self.points_A, 0)
            self.canvasB_magnifier = ImageMagnifier(self.canvasB, self.resized_imageB, self.points_B, 1)
            if self.is_aligning:
                self.canvasA_magnifier.toggle_magnifier(1)
                self.canvasB_magnifier.toggle_magnifier(1)
        else:
            print("图片未加载，跳过初始化放大镜")
    def clean_magnifier(self):
        """關閉並清除左右畫布的放大鏡元件。"""     
        if self.canvasA_magnifier:
            self.canvasA_magnifier.toggle_magnifier(0)
            self.canvasA_magnifier = None

        if self.canvasB_magnifier:
            self.canvasB_magnifier.toggle_magnifier(0)
            self.canvasB_magnifier = None
    def update_on_resize(self):
        """更新畫面顯示內容（視窗大小改變時觸發）。

        僅在畫布寬度確實發生變化時才重新渲染，避免不必要的效能消耗。
        根據對齊模式和放大鏡開關狀態決定是否初始化放大鏡。
        """
        # 强制更新所有待处理的几何变化，确保 winfo_width() 返回最新值
        self.root.update_idletasks()

        # 尺寸未变 不重复渲染
        old_canvas_width = self.canvasA_width
        new_canvas_width = self.canvasA.winfo_width()
        print(f"🔧 update_on_resize: old={old_canvas_width}, new={new_canvas_width}")
        if old_canvas_width == new_canvas_width:
            print(f"  ⏭️ 尺寸未變，跳過更新")
            return

        # 检查图片是否已加载
        if not hasattr(self, 'imageA') or not hasattr(self, 'imageB') or not self.imageA or not self.imageB:
            return

        # 直接更新图像显示，不修改原始坐标
        self.update_images()
        # 按开关控制放大镜
        if self.config.get("magnifier_switch") and self.is_aligning:
            self.init_magnifier()
        else:
            self.clean_magnifier()
    def update_images(self):
        """重新縮放並繪製熱力圖和 Layout 圖到畫布上。

        根據畫布當前尺寸計算縮放比例，保持原始寬高比。
        若處於對齊模式，會在圖片上繪製帶編號的對齊點標記。
        若有溫度標記框 (mark_rect_A/B)，會在非對齊模式下繪製。

        重要變數說明：
            self.imageA_scale: 熱力圖的縮放比例（畫布寬度 / 原始寬度）
            self.imageB_scale: Layout 圖的縮放比例
            self.canvasA_offset: 熱力圖在畫布上的偏移量 (offsetX, offsetY)
            self.canvasB_offset: Layout 圖在畫布上的偏移量
            self.resized_imageA: 縮放後的熱力圖 (PIL.Image)
            self.resized_imageB: 縮放後的 Layout 圖 (PIL.Image)
            self.tk_imageA: 畫布用的 PhotoImage 物件（需持有參考避免被 GC 回收）
            self.tk_imageB: 同上
        """
        # 檢查圖片是否存在，至少需要一張圖片
        if not hasattr(self, 'imageA') or not self.imageA:
            print("热力图未加载，跳过更新")
            return
            
        # 获取窗口的当前大小
        canvasA_width = self.canvasA.winfo_width()
        canvasA_height = self.canvasA.winfo_height()
        canvasB_width = self.canvasB.winfo_width()
        canvasB_height = self.canvasB.winfo_height()

        if canvasA_width <= 1 or canvasA_height <= 1:
            return

        # 计算每个图片的宽度和高度
        imageB_width = canvasB_width
        self.canvasA_width = canvasA_width

        # 矩形對齊模式：contain-fit（同時適應寬度和高度，圖片完整顯示不溢出）
        if self.is_rect_aligning:
            aspectA = self.imageA.height / self.imageA.width
            scale_by_w = canvasA_width / self.imageA.width
            scale_by_h = canvasA_height / self.imageA.height
            fit_scale = min(scale_by_w, scale_by_h)
            imageA_width = int(self.imageA.width * fit_scale)
            imageA_height = int(self.imageA.height * fit_scale)
        else:
            imageA_width = canvasA_width
            # 计算高度，保持原始宽高比
            aspectA = self.imageA.height / self.imageA.width
            imageA_height = int(imageA_width * aspectA)

        # 缩放热力图
        self.resized_imageA = self.imageA.resize((imageA_width, imageA_height), Image.LANCZOS)
        self.imageA_scale = imageA_width / self.imageA.width

        # 如果有Layout图，也进行缩放
        if hasattr(self, 'imageB') and self.imageB:
            aspectB = self.imageB.height / self.imageB.width
            imageB_height = int(imageB_width * aspectB)
            self.resized_imageB = self.imageB.resize((imageB_width, imageB_height), Image.LANCZOS)
            self.imageB_scale = imageB_width / self.imageB.width
        else:
            # 如果没有Layout图，创建一个空白图片
            imageB_height = imageA_height
            self.resized_imageB = Image.new('RGB', (imageB_width, imageB_height), color='white')
            self.imageB_scale = 1.0

        self.canvasA.delete("all")
        self.canvasB.delete("all")

        if self.is_aligning:
            imageB_np = self.to_numpy_image(self.resized_imageB)
            imageA_np = self.to_numpy_image(self.resized_imageA)

            # 将原始图像坐标转换为显示坐标
            if len(self.points_A) > 0:
                display_points_A = [[point[0] * self.imageA_scale, point[1] * self.imageA_scale] for point in self.points_A]
                imageA_np = draw_points_circle_ring_text(imageA_np, display_points_A)
            
            if len(self.points_B) > 0:
                display_points_B = [[point[0] * self.imageB_scale, point[1] * self.imageB_scale] for point in self.points_B]
                imageB_np = draw_points_circle_ring_text(imageB_np, display_points_B)
            
            self.resized_imageA = Image.fromarray(cv2.cvtColor(imageA_np, cv2.COLOR_BGR2RGB))
            self.resized_imageB = Image.fromarray(cv2.cvtColor(imageB_np, cv2.COLOR_BGR2RGB))

        self.tk_imageA = ImageTk.PhotoImage(self.resized_imageA)
        self.tk_imageB = ImageTk.PhotoImage(self.resized_imageB)
        # 清除画布上的旧图片
        # self.canvasA.delete("all")
        # self.canvasB.delete("all")
        # self.root.update_idletasks()  # 更新屏幕以确保显示效果

        # 更新背景图像位置和大小（基于中心点偏移 + 居中锚点）
        offsetA_x = (self.canvasA.winfo_width() - self.resized_imageA.width) // 2
        offsetA_y = (self.canvasA.winfo_height() - self.resized_imageA.height) // 2
        offsetB_x = (self.canvasB.winfo_width() - self.resized_imageB.width) // 2
        offsetB_y = (self.canvasB.winfo_height() - self.resized_imageB.height) // 2

        self.bg_imageA_id = self.canvasA.create_image(self.canvasA.winfo_width() // 2, self.canvasA.winfo_height() // 2, anchor=tk.CENTER, image=self.tk_imageA)
        self.bg_imageB_id = self.canvasB.create_image(self.canvasB.winfo_width() // 2, self.canvasB.winfo_height() // 2, anchor=tk.CENTER, image=self.tk_imageB)
        self.canvasA_offset = (offsetA_x, offsetA_y)
        self.canvasB_offset = (offsetB_x, offsetB_y)

        if not self.is_aligning and not self.is_rect_aligning and len(self.mark_rect_A) > 0:
            for itemA in self.mark_rect_A:
                draw_canvas_item(self.canvasA, itemA, self.imageA_scale, self.canvasA_offset, 0)
            # Layout 圖：傳入圖片邊界，框線裁切在圖片範圍內，名稱不消失
            clipB = (offsetB_x, offsetB_y,
                     offsetB_x + self.resized_imageB.width,
                     offsetB_y + self.resized_imageB.height)
            for itemB in self.mark_rect_B:
                draw_canvas_item(self.canvasB, itemB, self.imageB_scale, self.canvasB_offset, 1, clip_bounds=clipB)

        # if self.bg_imageA_id:
        #     self.canvasA.itemconfig(self.bg_imageA_id, image=self.tk_imageA)
        #     self.canvasA.coords(self.bg_imageA_id, 0, (canvasA_height - imageA_height) // 2)
        # else:
        #     self.bg_imageA_id = self.canvasA.create_image(0, (canvasA_height - imageA_height) // 2, anchor=tk.NW, image=self.tk_imageA)

        # if self.bg_imageB_id:
        #     self.canvasB.itemconfig(self.bg_imageB_id, image=self.tk_imageB)
        #     self.canvasB.coords(self.bg_imageB_id, 0, (canvasB_height - imageB_height) // 2)
        # else:
        #     self.bg_imageB_id = self.canvasB.create_image(0, (canvasB_height - imageB_height) // 2, anchor=tk.NW, image=self.tk_imageB)

        # # 在 Canvas 上绘制两张图片
        # self.canvasA.create_image(0, (canvasA_height - imageA_height) // 2, anchor=tk.NW, image=self.tk_imageA)
        # self.canvasB.create_image(0, (canvasB_height - imageB_height) // 2, anchor=tk.NW, image=self.tk_imageB)

        #缩放比  当前图片 / 原始图片
        self.canvasA.config(height=imageA_height)  # 重新设置 Canvas 的高度
        self.canvasB.config(height=imageB_height)  # 重新设置 Canvas 的高度

        # 矩形對齊模式：重繪 overlay
        if self.is_rect_aligning and self.rect_corners is not None:
            self._draw_rect_overlay()

    def on_resize(self, event):
        """視窗大小改變時的事件處理器。

        使用延遲更新機制（20ms debounce），避免拖曳縮放時頻繁觸發重繪。

        參數：
            event: Tkinter Configure 事件物件
        """
        # 每當視窗尺寸變化時，取消前一次的延遲更新，重新安排
        if self.resize_after:
            self.root.after_cancel(self.resize_after)
        self.resize_after = self.root.after(20, self.update_on_resize)
    def load_default_imgs(self, showTip = True):
        """載入預設圖片或從當前資料夾載入圖片。

        若已選擇資料夾，則重新掃描檔案並載入；否則嘗試載入預設路徑的圖片。

        參數：
            showTip (bool): 是否顯示載入結果的浮動通知
        """
        if self.current_folder_path:
            # 如果已经选择了文件夹，从文件夹中加载图片
            self.scan_folder_files()
            if showTip:
                show_toast(
                    title='加载成功',
                    message='已从当前文件夹加载图片',
                    duration=3000,
                    toast_type='success'
                )
        else:
            # 如果没有选择文件夹，尝试加载默认图片
            content = ""
            if os.path.isfile(Constants.imageA_default_path):  # 检查文件是否存在
                self.set_image(Constants.imageA_default_path, 0)
            else:
                content += Constants.imageA_default_path + " "
              
            if os.path.isfile(Constants.imageB_default_path):  # 检查文件是否存在
                self.set_image(Constants.imageB_default_path, 1)
            else:
                content += Constants.imageB_default_path + ""

            if showTip and content:
                show_toast(
                    title='文件不存在',
                    message= content + "文件不存在，\n请检查文件命名",
                    duration=5000,
                    toast_type='error'
                )
    def check_points_finish(self):
        """檢查兩側的對齊點是否已標記完成（各至少 3 個）。

        回傳：
            str: 若未完成，回傳錯誤訊息字串；若已完成，回傳空字串
        """
        content = ""
        if len(self.points_A) < 3:
        # if not os.path.isfile(Constants.imageA_point_path()):  # 检查文件是否存在
            content += "热力图未完成打点 "
          
        if len(self.points_B) < 3:
        #if not os.path.isfile(Constants.imageB_point_path()):  # 检查文件是否存在
           content += "Layout图未完成打点 "

        if content:
            show_toast(
                title='打点标记未完成',
                message= content + "\n请先对图片进行'打点标记'",
                duration=5000,
                toast_type='warning'
            )
        return content
    def update_points(self, clearAll = False):
        """更新畫布上的對齊點圓點顯示。

        將原始圖像座標經過縮放和偏移轉換為畫布座標後繪製。

        參數：
            clearAll (bool): True 表示清除所有對齊點顯示，False 表示重新繪製
        """
        if clearAll:
            self.canvasA.delete("points_A")
            self.canvasB.delete("points_B")
            return
        
        if not self.is_aligning:
            return
        
        radius = 4
        self.canvasA.delete("points_A")
        
        # 将原始图像坐标转换为canvas坐标
        offAx, offAy = getattr(self, 'canvasA_offset', (0, 0))
        for point in self.points_A:
            # 原始图像坐标 -> 显示坐标
            display_x = point[0] * self.imageA_scale
            display_y = point[1] * self.imageA_scale
            
            # 显示坐标 -> canvas坐标（中心偏移）
            canvas_x = display_x + offAx
            canvas_y = display_y + offAy
            
            y0 = min(canvas_y - radius, self.canvasA.winfo_height())
            y1 = min(canvas_y + radius, self.canvasA.winfo_height())
            self.canvasA.create_oval(canvas_x - radius, y0, canvas_x + radius, y1, fill="black", tags="points_A")

        self.canvasB.delete("points_B")
        offBx, offBy = getattr(self, 'canvasB_offset', (0, 0))
        for point in self.points_B:
            # 原始图像坐标 -> 显示坐标
            display_x = point[0] * self.imageB_scale
            display_y = point[1] * self.imageB_scale
            
            # 显示坐标 -> canvas坐标（中心偏移）
            canvas_x = display_x + offBx
            canvas_y = display_y + offBy
            
            y0 = min(canvas_y - radius, self.canvasB.winfo_height())
            y1 = min(canvas_y + radius, self.canvasB.winfo_height())
            self.canvasB.create_oval(canvas_x - radius, y0, canvas_x + radius, y1, fill="red", tags="points_B")
    def save_points_json(self):
        """將對齊點數據儲存為 JSON 格式檔案。

        檔案儲存在 {資料夾}/points/{熱力圖名}_{Layout圖名}.json，
        使用原始圖像座標（非畫布座標），同時記錄圖像尺寸和時間戳記。
        需要兩側各至少 3 個對齊點才會儲存。
        """
        try:
            if len(self.points_A) >= 3 and len(self.points_B) >= 3:
                # 获取原始图像尺寸
                aW, aH = self.imageA.size
                bW, bH = self.imageB.size
                
                # 统一转为可序列化的list（兼容list/ndarray）
                import numpy as _np
                points_A_list = _np.asarray(self.points_A).tolist()
                points_B_list = _np.asarray(self.points_B).tolist()

                points_data = {
                    'points_A': points_A_list,
                    'points_B': points_B_list,
                    'image_A_size': [aW, aH],
                    'image_B_size': [bW, bH],
                    'alignment_type': self.alignment_type,
                    'timestamp': datetime.now().isoformat()
                }
                
                if self.current_folder_path:
                    points_dir = os.path.join(self.current_folder_path, "points")
                    if not os.path.exists(points_dir):
                        os.makedirs(points_dir)
                    # 使用规范命名：热力图文件名 + '_' + Layout文件名 + '.json'
                    heat_filename = self.current_files.get("heat", "")
                    layout_filename = self.current_files.get("layout", "")
                    if not heat_filename or not layout_filename:
                        print("保存打点数据失败：缺少热力图或Layout图文件名")
                        return
                    heat_name = os.path.splitext(heat_filename)[0]
                    layout_name = os.path.splitext(layout_filename)[0]
                    points_file = os.path.join(points_dir, f"{heat_name}_{layout_name}.json")
                    print(f"保存打点数据到: {points_file}")
                    
                    with open(points_file, 'w', encoding='utf-8') as f:
                        json.dump(points_data, f, indent=2, ensure_ascii=False)
                    
                    print("打点数据已保存为JSON格式")
                else:
                    print("没有当前文件夹路径，无法保存打点数据")
            else:
                print(f"打点数量不足，无法保存（A: {len(self.points_A)}, B: {len(self.points_B)}）")
        except Exception as e:
            print(f"保存打点数据失败: {e}")
            import traceback
            traceback.print_exc()

    def save_points_csv(self):
        """將對齊點數據儲存為 CSV 格式檔案（舊版格式，已逐步被 JSON 取代）。

        CSV 檔案第一行為畫布尺寸 [w, h]，後續行為對齊點座標。
        檔案命名格式：{熱力圖名}_{Layout圖名}_imageA.csv / _imageB.csv
        """
        aW, aH = self.resized_imageA.size
        points_A_save = np.vstack([np.array([[aW, aH]], dtype='float32'), self.points_A])
        # 检查 points_A 的每一行数据是否符合条件
        A_save = True
        # for i in range(0, len(self.points_A)):  # 从第二行开始检查
        #     if self.points_A[i, 0] > aW or self.points_A[i, 1] > aH:
        #         A_save = False
        #         print("Data violation found, returning without saving.")
        #         break 
        if A_save:
            # 使用正确的路径构建方式
            if self.current_folder_path:
                points_dir = os.path.join(self.current_folder_path, "points")
                if not os.path.exists(points_dir):
                    os.makedirs(points_dir)
                
                # 使用新的文件名格式：{热力图文件名}_{Layout图文件名}_imageA.csv
                heat_filename = self.current_files.get("heat", "")
                layout_filename = self.current_files.get("layout", "")
                
                if heat_filename and layout_filename:
                    # 去掉文件扩展名
                    heat_name = os.path.splitext(heat_filename)[0]
                    layout_name = os.path.splitext(layout_filename)[0]
                    
                    # 构建新的点位文件名
                    imageA_points_filename = f"{heat_name}_{layout_name}_imageA.csv"
                    imageA_points_path = os.path.join(points_dir, imageA_points_filename)
                    print(f"save_points_csv: 保存热力图点位到 {imageA_points_path}")
                    np.savetxt(imageA_points_path, points_A_save, delimiter=',', fmt='%d')
                else:
                    print("save_points_csv: 缺少热力图或Layout图文件名，无法保存点位数据")
            else:
                np.savetxt(Constants.imageA_point_path(), points_A_save, delimiter=',', fmt='%d')

        bW, bH = self.resized_imageB.size
        points_B_save = np.vstack([np.array([[bW, bH]], dtype='float32'), self.points_B])
        B_save = True
        # for i in range(0, len(self.points_B)):  # 从第二行开始检查
        #     if self.points_B[i, 0] > bW or self.points_B[i, 1] > bH:
        #         B_save = False
        #         print("Data violation found, returning without saving.")
        #         break 
        if B_save:
            # 使用正确的路径构建方式
            if self.current_folder_path:
                points_dir = os.path.join(self.current_folder_path, "points")
                if not os.path.exists(points_dir):
                    os.makedirs(points_dir)
                
                # 使用新的文件名格式：{热力图文件名}_{Layout图文件名}_imageB.csv
                heat_filename = self.current_files.get("heat", "")
                layout_filename = self.current_files.get("layout", "")
                
                if heat_filename and layout_filename:
                    # 去掉文件扩展名
                    heat_name = os.path.splitext(heat_filename)[0]
                    layout_name = os.path.splitext(layout_filename)[0]
                    
                    # 构建新的点位文件名
                    imageB_points_filename = f"{heat_name}_{layout_name}_imageB.csv"
                    imageB_points_path = os.path.join(points_dir, imageB_points_filename)
                    print(f"save_points_csv: 保存Layout图点位到 {imageB_points_path}")
                    np.savetxt(imageB_points_path, points_B_save, delimiter=',', fmt='%d')
                else:
                    print("save_points_csv: 缺少热力图或Layout图文件名，无法保存点位数据")
            else:
                np.savetxt(Constants.imageB_point_path(), points_B_save, delimiter=',', fmt='%d')
    def get_points(self, points_path, canvas):
        """從 CSV 檔案讀取對齊點數據，並根據當前畫布尺寸進行縮放。

        參數：
            points_path (str): 對齊點 CSV 檔案路徑
            canvas (tk.Canvas): 對應的畫布元件（用於取得當前寬度）

        回傳：
            list: 縮放後的對齊點座標列表，若檔案不存在或數據不足則回傳空列表
        """
         # 如果文件不存在，返回空数组
        if not os.path.exists(points_path):
            return []
        data = np.loadtxt(points_path, delimiter=',', dtype=np.float32)
        if data.shape[0] < 4:
            return []
        w, h = data[0]  # 第一行代表宽（w）和高（h）
        points = data[1:]  # 剩下的3行是坐标点
        scale = canvas.winfo_width() / w   # 当窗口打开时，不是保存时的窗口大小了
        print("get_points -> ", canvas.winfo_width(), w, scale, points * scale)
        return points * scale
    def clear_point_file(self):
        """刪除當前檔案組合對應的對齊點 CSV 檔案。"""
        if self.current_folder_path:
            points_dir = os.path.join(self.current_folder_path, "points")
            
            # 使用新的文件名格式
            heat_filename = self.current_files.get("heat", "")
            layout_filename = self.current_files.get("layout", "")
            
            if heat_filename and layout_filename:
                # 去掉文件扩展名
                heat_name = os.path.splitext(heat_filename)[0]
                layout_name = os.path.splitext(layout_filename)[0]
                
                # 构建新的点位文件名
                imageA_points_filename = f"{heat_name}_{layout_name}_imageA.csv"
                imageB_points_filename = f"{heat_name}_{layout_name}_imageB.csv"
                
                imageA_points_path = os.path.join(points_dir, imageA_points_filename)
                imageB_points_path = os.path.join(points_dir, imageB_points_filename)
                
                print(f"clear_point_file: 清除点位文件 - {imageA_points_filename}, {imageB_points_filename}")
                self.remove_file(imageA_points_path)
                self.remove_file(imageB_points_path)
            else:
                print("clear_point_file: 缺少热力图或Layout图文件名，无法清除点位文件")
        else:
            self.remove_file(Constants.imageA_point_path())
            self.remove_file(Constants.imageB_point_path())
    def remove_file(self, file_path):
        """安全刪除指定檔案，捕獲常見例外（檔案不存在、權限不足等）。

        參數：
            file_path (str): 要刪除的檔案完整路徑
        """
        try:
            os.remove(file_path)
            # print(f"{file_path} 已成功删除。")
        except FileNotFoundError:
            print(f"文件 {file_path} 未找到。")
        except PermissionError:
            print(f"没有权限删除文件 {file_path}。")
        except Exception as e:
            print(f"删除文件时发生错误: {e}")
    def unbind_point_event(self, canvas):
        # 绑定鼠标事件
        canvas.unbind("<Button-1>")
        canvas.unbind("<Button-3>")
    def bind_point_event(self, canvas, index):
        canvas.bind("<Button-1>", lambda event: self.point_mouse_click(event, index)) # 左键
        canvas.bind("<Button-3>", lambda event: self.point_mouse_click(event, index))  # 右键
        # self.canvasA.bind("<Motion>", self.point_mouse_move)     # 鼠标移动

    def get_click_point(self, circles, x, y):
        if self.config.get("circle_switch"):
            ret = find_circle_containing_point(circles, x, y)
            if ret:
                return [ret[0], ret[1]]
        
        return [x, y]

    def point_mouse_click(self, event, index):
        """处理打点点击事件 - 将canvas坐标转换为原始图像坐标"""
        x, y = event.x, event.y
        range = 16
        
        if index == 0:
            points = self.points_A.copy()
            offx, offy = getattr(self, 'canvasA_offset', (0, 0))
            scale = self.imageA_scale
            recognize_circles = self.recognize_circle_A 
        else:
            points = self.points_B.copy()
            offx, offy = getattr(self, 'canvasB_offset', (0, 0))
            scale = self.imageB_scale
            recognize_circles = self.recognize_circle_B

        # canvas坐标 -> 显示坐标
        display_x = x - offx
        display_y = y - offy
        
        # 显示坐标 -> 原始图像坐标
        original_x = display_x / scale
        original_y = display_y / scale
        
        if event.num == 1:  # 左键点击
            # 允许更多对齐点（提升精度），上限设为8个
            MAX_POINTS = 8
            if len(points) >= MAX_POINTS:
                messagebox.showinfo("提示", f"最多标记{MAX_POINTS}个点")
                return
            # 使用原始图像坐标
            points.append([original_x, original_y])
            self.pont_marked = True
            print(f"左键点击: canvas({x}, {y}) -> 原始图像({original_x:.1f}, {original_y:.1f})")
        elif event.num == 3:  # 右键点击
            print("point_mouse_click1 -> ", points)
            # 在原始图像坐标中查找要删除的点
            points = [[cx, cy] for cx, cy in points if not (original_x - range/scale <= cx <= original_x + range/scale and original_y - range/scale <= cy <= original_y + range/scale)]
            self.pont_marked = True
            print(f"右键点击: canvas({x}, {y}) -> 原始图像({original_x:.1f}, {original_y:.1f})")

        print("point_mouse_click2 -> ", points)
        if index == 0:
            self.points_A = points
        else:
            self.points_B = points

        self.update_images()
        if self.config.get("magnifier_switch") and self.is_aligning:
            self.init_magnifier()
        else:
            self.clean_magnifier()
        # self.update_points()
        # self.update_magnifier_point()
    def start_point_mark(self):
        if self.is_aligning:
            if self.check_points_finish():  # 检查是否打点完成（各≥3）
                return
            # 检查两侧点数是否一致
            if len(self.points_A) != len(self.points_B):
                show_toast(
                    title='打点数量不匹配',
                    message=f'A侧: {len(self.points_A)} 个, B侧: {len(self.points_B)} 个\n请保证两侧点数相同且≥3',
                    duration=5000,
                    toast_type='warning'
                )
                return
            # 先尝试构建转换器，失败则提示并留在打点模式
            try:
                temp_transformer = PointTransformer(self.points_A, self.points_B)
            except Exception as e:
                show_toast(
                    title='对齐失败',
                    message=f'点位异常：{e}\n请检查两侧点的对应关系与数量',
                    duration=6000,
                    toast_type='error'
                )
                return
            
            self.is_aligning = False

            # 重新打点 清除编辑框
            if self.pont_marked:
                self.mark_rect_A = []
                self.mark_rect_B = []
                self.pont_marked = False

            # self.update_points(True)
            # 保存结果
            self.unbind_point_event(self.canvasA)
            self.unbind_point_event(self.canvasB)
            self.save_points_json()
            self.update_images()
            # 保存后立刻刷新对齐按钮可见性
            self.update_align_buttons_visibility()

            # 对齐工具类
            self.point_transformer = temp_transformer

            # 放大镜：结束打点后根据开关关闭
            self.clean_magnifier()

            # self.start_margin()
            self.alignment_type = 'multi_point'
            self.align_button.config(text="多点对齐")  # 切换为结束状态
            # 恢復矩形對齊按鈕
            self.rect_align_button.config(state=tk.NORMAL)
            # 隐藏清除对齐点按钮
            self.clear_heat_points_button.grid_forget()
            self.clear_layout_points_button.grid_forget()
            # 隐藏Layout图下方的按钮框架
            self.bottom_buttons_frame_B.grid_forget()
            # 显示原来的按钮
            self.margin_before_button.grid(row=0, column=0, padx=5)
            self.margin_after_button.grid(row=0, column=1, padx=5)

        else:
            self.is_aligning = True
            self.bind_point_event(self.canvasA, 0)
            self.bind_point_event(self.canvasB, 1)
            # self.update_points()
            self.update_images()

            if self.config.get("magnifier_switch"):
                self.init_magnifier()

            # 禁用矩形對齊按鈕
            self.rect_align_button.config(state=tk.DISABLED)

            # self.points_B = np.array(np.loadtxt(Constants.imageB_point_path(), delimiter=','), dtype=np.float32) * self.window_scale
            # 读取本地文件
            self.margin_before_button.grid_forget()
            self.margin_after_button.grid_forget()
            # 显示清除对齐点按钮
            self.clear_heat_points_button.grid(row=0, column=0, padx=5, pady=5)
            self.clear_layout_points_button.grid(row=0, column=0, padx=5, pady=5)
            # 显示Layout图下方的按钮框架
            self.bottom_buttons_frame_B.grid(row=2, column=2)
            self.align_button.config(text="对齐图像结束")  # 切换为开始状态
    
    def clear_heat_points(self):
        """清除热力图的对齐点"""
        try:
            # 清除热力图的对齐点数据
            self.points_A = []
            self.mark_rect_A = []
            
            # 清除对应的点位文件
            if self.current_folder_path:
                points_dir = os.path.join(self.current_folder_path, "points")
                heat_filename = self.current_files.get("heat", "")
                pcb_filename = self.current_files.get("pcb", "")
                
                if heat_filename and pcb_filename:
                    # 去掉文件扩展名
                    heat_name = os.path.splitext(heat_filename)[0]
                    pcb_name = os.path.splitext(pcb_filename)[0]
                    
                    # 构建点位文件名
                    imageA_points_filename = f"{heat_name}_{pcb_name}_imageA.csv"
                    imageA_points_path = os.path.join(points_dir, imageA_points_filename)
                    
                    print(f"clear_heat_points: 删除点位文件 {imageA_points_path}")
                    self.remove_file(imageA_points_path)
            
            # 清除画布上的标记
            self.canvasA.delete("all")
            if self.bg_imageA_id:
                self.canvasA.delete(self.bg_imageA_id)
                self.bg_imageA_id = None
            
            # 重新显示图片
            if self.imageA:
                self.update_images()
            
            # 重新初始化放大镜（严格按开关）
            if self.is_aligning and self.config.get("magnifier_switch"):
                self.init_magnifier()
            else:
                self.clean_magnifier()
            
            # 显示成功消息
            show_toast(
                title='清除成功',
                message='已清除热力图对齐点',
                duration=3000,
                toast_type='success'
            )
            print("已清除热力图对齐点")
            
        except Exception as e:
            print(f"清除热力图对齐点时出错: {e}")
            from tkinter import messagebox
            messagebox.showerror("错误", f"清除热力图对齐点失败: {e}")
    
    def clear_layout_points(self):
        """清除Layout图的对齐点"""
        try:
            # 清除Layout图的对齐点数据
            self.points_B = []
            self.mark_rect_B = []
            
            # 清除对应的点位文件
            if self.current_folder_path:
                points_dir = os.path.join(self.current_folder_path, "points")
                heat_filename = self.current_files.get("heat", "")
                layout_filename = self.current_files.get("layout", "")
                
                if heat_filename and layout_filename:
                    # 去掉文件扩展名
                    heat_name = os.path.splitext(heat_filename)[0]
                    layout_name = os.path.splitext(layout_filename)[0]
                    
                    # 构建点位文件名
                    imageB_points_filename = f"{heat_name}_{layout_name}_imageB.csv"
                    imageB_points_path = os.path.join(points_dir, imageB_points_filename)
                    
                    print(f"clear_layout_points: 删除点位文件 {imageB_points_path}")
                    self.remove_file(imageB_points_path)
            
            # 清除画布上的标记
            self.canvasB.delete("all")
            if self.bg_imageB_id:
                self.canvasB.delete(self.bg_imageB_id)
                self.bg_imageB_id = None
            
            # 重新显示图片
            if self.imageB:
                self.update_images()
            
            # 重新初始化放大镜（如果在打点模式下且放大镜开关开启）
            if self.is_aligning and self.config.get("magnifier_switch"):
                self.init_magnifier()
            
            # 显示成功消息
            show_toast(
                title='清除成功',
                message='已清除Layout图对齐点',
                duration=3000,
                toast_type='success'
            )
            print("已清除Layout图对齐点")
            
        except Exception as e:
            print(f"清除Layout图对齐点时出错: {e}")
            from tkinter import messagebox
            messagebox.showerror("错误", f"清除Layout图对齐点失败: {e}")
    
    def load_points(self):
        """加载点位数据，现在从选择的文件夹中加载"""
        print(f"load_points: current_folder_path = {self.current_folder_path}")
        
        if self.current_folder_path:
            # 如果已经选择了文件夹，从文件夹中加载点位数据
            points_dir = os.path.join(self.current_folder_path, "points")
            print(f"load_points: points_dir = {points_dir}, exists = {os.path.exists(points_dir)}")

            if os.path.exists(points_dir):
                # 使用当前选择的热力图和Layout图文件名构建点位文件名（JSON）
                heat_filename = self.current_files.get("heat", "")
                layout_filename = self.current_files.get("layout", "")
                
                if heat_filename and layout_filename:
                    # 去掉文件扩展名
                    heat_name = os.path.splitext(heat_filename)[0]
                    layout_name = os.path.splitext(layout_filename)[0]
                    
                    json_points_path = os.path.join(points_dir, f"{heat_name}_{layout_name}.json")
                    print(f"load_points: 尝试加载 {json_points_path}, exists = {os.path.exists(json_points_path)}")
                    
                    if os.path.exists(json_points_path):
                        with open(json_points_path, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                        self.points_A = data.get('points_A', [])
                        self.points_B = data.get('points_B', [])
                        self.alignment_type = data.get('alignment_type', 'multi_point')
                        print(f"load_points: loaded points_A = {self.points_A}")
                        print(f"load_points: loaded points_B = {self.points_B}")
                        print(f"load_points: alignment_type = {self.alignment_type}")

                        # 若是矩形對齊，恢復 rect_corners
                        if self.alignment_type == 'rect' and len(self.points_A) == 4:
                            self.rect_corners = [tuple(p) for p in self.points_A]

                        if hasattr(self, 'imageA') and hasattr(self, 'imageB') and self.imageA and self.imageB:
                            self.init_point_transformer()
                        # 加载完点位后，立即刷新按钮可见性
                        self.update_align_buttons_visibility()
                    else:
                        print("load_points: 未找到json点位文件")
                else:
                    print("load_points: 缺少热力图或Layout图文件名，无法加载点位数据")
            else:
                print("load_points: points目录不存在")
        else:
            # 如果没有选择文件夹，尝试加载user_data/A/points下的默认点位数据
            print("load_points: 没有选择文件夹，尝试加载默认点位数据")
            default_imageA_path = "user_data/A/points/imageA.jpg_points.csv"
            default_imageB_path = "user_data/A/points/imageB.jpg_points.csv"
            
            print(f"load_points: 尝试加载 {default_imageA_path}, exists = {os.path.exists(default_imageA_path)}")
            print(f"load_points: 尝试加载 {default_imageB_path}, exists = {os.path.exists(default_imageB_path)}")
            
            self.points_A = self.get_points(default_imageA_path, self.canvasA)
            self.points_B = self.get_points(default_imageB_path, self.canvasB)
            
            print(f"load_points: 默认加载 points_A = {self.points_A}")
            print(f"load_points: 默认加载 points_B = {self.points_B}")
            
            if len(self.points_A) > 0 and len(self.points_B) > 0:
                self.init_point_transformer()

    def init_point_transformer(self):
        """初始化点转换器"""
        if len(self.points_A) > 0:
            matched = (self.alignment_type == 'rect')
            self.point_transformer = PointTransformer(self.points_A, self.points_B, matched=matched)
    
    def clear_and_reload_points(self):
        """清空当前对齐点数据并重新加载对应文件的点位数据"""
        try:
            print("clear_and_reload_points: 清空当前对齐点数据")
            
            # 清空当前的对齐点数据
            self.points_A = []
            self.points_B = []
            self.mark_rect_A = []
            self.mark_rect_B = []
            self.point_transformer = None
            
            # 清空画布上的标记
            if hasattr(self, 'canvasA'):
                self.canvasA.delete("all")
                if self.bg_imageA_id:
                    self.canvasA.delete(self.bg_imageA_id)
                    self.bg_imageA_id = None
            
            if hasattr(self, 'canvasB'):
                self.canvasB.delete("all")
                if self.bg_imageB_id:
                    self.canvasB.delete(self.bg_imageB_id)
                    self.bg_imageB_id = None
            
            # 重新显示图片（不显示对齐点）
            if hasattr(self, 'imageA') and hasattr(self, 'imageB') and self.imageA and self.imageB:
                self.update_images()
            
            # 尝试加载新文件组合对应的点位数据
            self.load_points()
            
            print("clear_and_reload_points: 完成清空和重新加载")
            
        except Exception as e:
            print(f"clear_and_reload_points 出错: {e}")
            import traceback
            traceback.print_exc()

    def to_mark_rect_B(self, itemA):
        # 初始化一个空字典
        ret = {}
        # 确保有point_transformer对象
        if self.point_transformer:
            # 对特定字段进行转换
            ret["x1"], ret["y1"] = self.point_transformer.A_2_oriB(itemA.get("x1"), itemA.get("y1"))
            ret["x2"], ret["y2"] = self.point_transformer.A_2_oriB(itemA.get("x2"), itemA.get("y2"))
            ret["cx"], ret["cy"] = self.point_transformer.A_2_oriB(itemA.get("cx"), itemA.get("cy"))

        # 复制 itemA 中其他字段到 ret 字典
        for key, value in itemA.items():
            if key not in ret:
                ret[key] = value

        return ret
    def on_close_editor(self, mark_rect_A, add_new_count, delete_new_count, modify_origin_set):
        self.edit_log["add_new_mark"][1] += add_new_count
        self.edit_log["delete_origin_mark"][1] += delete_new_count
        self.edit_log["modify_origin_mark"][1].update(modify_origin_set)

        print("modify_origin_set -------->>> ", self.edit_log, len(modify_origin_set))

        if len(mark_rect_A) > 0:
            self.mark_rect_A = mark_rect_A

            ret = []
            for itemA in mark_rect_A: 
                ret.append(self.to_mark_rect_B(itemA))
            self.mark_rect_B = ret

            self.update_images()
        
        # 清空EditorCanvas实例引用
        self.editor_canvas = None
    def on_template_confirm(self, dialog_result):
        min_temp, max_temp = dialog_result.get("min_temp"), dialog_result.get("max_temp")
        # min_width 和 min_height 已移除（未實際使用）

        # 获取新的PCB参数
        p_w = dialog_result.get("p_w", 237)
        p_h = dialog_result.get("p_h", 194)
        p_origin = dialog_result.get("p_origin", "左下")
        p_origin_offset_x = dialog_result.get("p_origin_offset_x", 0)
        p_origin_offset_y = dialog_result.get("p_origin_offset_y", 0)
        c_padding_left = dialog_result.get("c_padding_left", 0)
        c_padding_top = dialog_result.get("c_padding_top", 0)
        c_padding_right = dialog_result.get("c_padding_right", 0)
        c_padding_bottom = dialog_result.get("c_padding_bottom", 0)
        
        # 检查并初始化point_transformer
        if self.point_transformer is None:
            if len(self.points_A) > 0 and len(self.points_B) > 0:
                self.init_point_transformer()
            else:
                # messagebox.showwarning("警告", "请先进行图像对齐")
                show_toast(
                    title='警告',
                    message= "请先进行图像对齐",
                    duration=5000,
                    toast_type='warning'
                )
                return
        
        # 检查tempALoader是否存在
        if not hasattr(self, 'tempALoader') or self.tempALoader is None:
            print("警告：tempALoader不存在，请先加载温度数据文件")
            messagebox.showwarning("警告", "请先加载温度数据文件")
            return
        
        # 检查Layout数据是否存在
        if not hasattr(self, 'layout_data') or self.layout_data is None:
            print("警告：Layout数据不存在，请先加载Layout数据文件")
            messagebox.showwarning("警告", "请先加载Layout数据文件")
            return
        
        # 检查Layout数据是否为空列表
        if isinstance(self.layout_data, list) and len(self.layout_data) == 0:
            print("警告：Layout数据为空，请检查Layout数据文件")
            messagebox.showwarning("警告", "Layout数据为空，请检查Layout数据文件")
            return
        
        # 使用新的Layout查询方法
        try:
            # 检查必要的数据是否存在
            print("=== 开始Layout温度查询 ===")
            print(f"Layout数据: {self.layout_data is not None and len(self.layout_data) if self.layout_data else 0} 个元器件")
            print(f"温度数据: {self.tempALoader.get_tempA().shape if self.tempALoader and self.tempALoader.get_tempA() is not None else 'None'}")
            print(f"点转换器: {self.point_transformer is not None}")
            print(f"Layout图像: {self.imageB.size if self.imageB else 'None'}")
            
            if self.layout_data is None or len(self.layout_data) == 0:
                raise Exception("Layout数据为空，请先加载Layout数据文件")
            
            if self.tempALoader is None or self.tempALoader.get_tempA() is None:
                raise Exception("温度数据为空，请先加载温度数据文件")
            
            if self.point_transformer is None:
                raise Exception("点转换器未初始化，请先完成图像对齐")
            
            # 使用优化版的温度查询
            layout_query = LayoutTemperatureQueryOptimized(
                layout_data=self.layout_data,
                temp_data=self.tempALoader.get_tempA().copy(),
                point_transformer=self.point_transformer,
                p_w=p_w,
                p_h=p_h,
                p_origin=p_origin,
                p_origin_offset_x=p_origin_offset_x,
                p_origin_offset_y=p_origin_offset_y,
                c_padding_left=c_padding_left,
                c_padding_top=c_padding_top,
                c_padding_right=c_padding_right,
                c_padding_bottom=c_padding_bottom,
                layout_image=self.imageB  # 传递Layout图像
            )
            
            # 执行智能过滤版温度查询
            self.mark_rect_A, self.mark_rect_B = layout_query.query_temperature_by_layout_smart_filter(min_temp, max_temp)
            # 保存原始辨識結果（用於 EditorCanvas 回到起點跨 session 恢復）
            self.origin_mark_rect_A = copy.deepcopy(self.mark_rect_A)

            print(f"智能过滤版Layout查询完成，找到 {len(self.mark_rect_A)} 个高温元器件")
            
        except Exception as e:
            print(f"Layout查询出错: {e}")
            import traceback
            traceback.print_exc()
            
            # 不使用YOLO回退，强制使用Layout方法
            print("Layout查询失败，请检查数据完整性")
            messagebox.showerror("错误", f"Layout温度查询失败: {e}\n请检查：\n1. Layout数据是否正确加载\n2. 温度数据是否正确加载\n3. 图像对齐是否完成")
            
            # 清空结果
            self.mark_rect_A = []
            self.mark_rect_B = []

        self.update_images()
        

        # temp_imageA = Image.fromarray(cv2.cvtColor(self.imageA_cv_export, cv2.COLOR_BGR2RGB))
        # cv2.imshow("imageA mark", cv2.resize(imageA_cv, (1024, 768)))
        # cv2.imshow("imageA mark", cv2.resize(imageA_cv, (1024, 768)))
        # EditorCanvas(root, image=Image.open("imageA.jpg"), mark_rect=mark_rect, on_close_callback=on_window_close)
        self.edit_log = copy.deepcopy(DEFAULT_EDIT_LOG)
        # self.edit_log["modify_origin_mark"][1].clear()
        self.edit_log["origin_mark"][1] = len(self.mark_rect_A)
        # EditorCanvas(self.root, self.imageA, self.mark_rect_A, on_close_callback=self.on_close_editor)

    def open_template_dialog(self):
        print("点击温度过滤按钮，开始创建对话框...")
        try:
            print(f"当前文件夹路径: {self.current_folder_path}")
            templateDialog = TemplateDialog(self.root, self.template_filter_button, self.on_template_confirm, self.current_folder_path)
            print("TemplateDialog创建成功，准备打开对话框...")
            
            # 在打开对话框前，同步文件信息到温度配置管理器
            print("同步文件信息到温度配置管理器...")
            self.update_temp_config_files()
            
            templateDialog.open()
            print("对话框打开完成")
        except Exception as e:
            print(f"创建或打开对话框时出错: {e}")
            import traceback
            traceback.print_exc()
        # if self.min_temp:
        #     templateDialog.open(self.min_temp, self.max_temp, self.min_width, self.min_height, self.max_ratio, self.auto_reduce)
        # else:    
        #     templateDialog.open()
    def export_excel(self):
        if self.mark_rect_A:
            # 确定输出目录路径
            if self.current_folder_path:
                output_dir = os.path.join(self.current_folder_path, "output")
            else:
                output_dir = "output"
            
            # 创建输出目录
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # 创建一个新的 Excel 工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "温度报告"
            # 添加标题行（比照 EditorCanvas Treeview 三欄）
            ws.append(["點位名稱", "描述", "溫度"])
            # 将 rect_arr 中的数据写入到 Excel 文件
            for item in self.mark_rect_A:
                max_temp = item.get("max_temp", 0)
                ws.append([item.get("name", ""), item.get("description", ""), f"{max_temp:.1f}°C"])
               
            # 保存Excel文件到当前文件夹的output目录，如果文件被占用则自动重命名
            excel_path = self.get_available_excel_path(output_dir, "report.xlsx")
            wb.save(excel_path)

            # 保存图片到当前文件夹的output目录
            imageA_input = cv2.cvtColor(np.array(self.imageA), cv2.COLOR_RGB2BGR)
            imageA_output = draw_numpy_image_item(imageA_input, self.mark_rect_A)
            image_path = os.path.join(output_dir, "A.jpg")
            Image.fromarray(cv2.cvtColor(imageA_output, cv2.COLOR_BGR2RGB)).save(image_path, quality=100)

            #输出日志
            self.edit_log["final_mark"][1] = len(self.mark_rect_A)
            self.save_log_file()

            show_toast(
                title='导出成功',
                message= f"导出报告成功，报告位于 {excel_path}",
                duration=5000,
                toast_type='success'
            )
        else:
            show_toast(
                title='导出失败',
                message= "请先进行'温度过滤'， 找出温度区域",
                duration=5000,
                toast_type='error'
            )
    
    def get_available_excel_path(self, output_dir, base_filename):
        """
        获取可用的Excel文件路径，如果文件被占用则自动重命名
        
        Args:
            output_dir: 输出目录
            base_filename: 基础文件名（如 "report.xlsx"）
            
        Returns:
            str: 可用的文件路径
        """
        import os
        
        # 分离文件名和扩展名
        name, ext = os.path.splitext(base_filename)
        
        # 尝试原始文件名
        original_path = os.path.join(output_dir, base_filename)
        
        # 如果文件不存在，直接返回原始路径
        if not os.path.exists(original_path):
            return original_path
        
        # 如果文件存在，尝试重命名
        counter = 1
        while True:
            new_filename = f"{name}{counter}{ext}"
            new_path = os.path.join(output_dir, new_filename)
            
            if not os.path.exists(new_path):
                print(f"文件 {base_filename} 已存在，使用新文件名: {new_filename}")
                return new_path
            
            counter += 1
            
            # 防止无限循环，最多尝试100次
            if counter > 100:
                print(f"警告：无法找到可用的文件名，使用时间戳")
                import time
                timestamp = int(time.time())
                timestamp_filename = f"{name}_{timestamp}{ext}"
                return os.path.join(output_dir, timestamp_filename)

        # print("xx--> export_excel")

    def write_test_report(self):
        """開啟 Sheet 名稱選擇對話框，將熱力圖影像和元器件列表寫入測試報告 .xlsx。"""
        # 檢查是否有標記數據
        if not self.mark_rect_A:
            show_toast(
                title='寫入失敗',
                message="請先進行'溫度過濾'，找出溫度區域",
                duration=5000,
                toast_type='error'
            )
            return

        # 檢查測試報告檔案
        test_report_file = self.current_files.get("testReport")
        if not test_report_file or not self.current_folder_path:
            show_toast(
                title='寫入失敗',
                message="未偵測到測試報告檔案",
                duration=5000,
                toast_type='error'
            )
            return

        # 開啟 Sheet 名稱選擇對話框
        dialog = ExportReportDialog(self.root, self.test_report_button, self._do_write_test_report)
        dialog.open()

    def _do_write_test_report(self, sheet_name):
        """實際執行寫入測試報告的邏輯（由對話框 callback 呼叫）。

        Args:
            sheet_name (str): 使用者選擇的 Sheet 名稱
        """
        test_report_file = self.current_files.get("testReport")
        report_path = os.path.join(self.current_folder_path, test_report_file)

        try:
            wb = openpyxl.load_workbook(report_path)

            # 若 sheet 名稱重複，加上數字後綴
            final_name = sheet_name
            existing_names = wb.sheetnames
            if final_name in existing_names:
                counter = 1
                while f"{sheet_name}_{counter}" in existing_names:
                    counter += 1
                final_name = f"{sheet_name}_{counter}"

            ws = wb.create_sheet(title=final_name)

            # 寫入表頭於 A1（完全比照 EditorCanvas 左側 Treeview 的欄位）
            from openpyxl.styles import Font
            bold_font = Font(bold=True)
            ws.cell(row=1, column=1, value="點位名稱").font = bold_font
            ws.cell(row=1, column=2, value="描述").font = bold_font
            ws.cell(row=1, column=3, value="溫度").font = bold_font

            # 寫入元器件資料（格式與 EditorCanvas Treeview 一致）
            for i, item in enumerate(self.mark_rect_A):
                row = 2 + i
                ws.cell(row=row, column=1, value=item.get("name", ""))
                ws.cell(row=row, column=2, value=item.get("description", ""))
                max_temp = item.get("max_temp", 0)
                ws.cell(row=row, column=3, value=f"{max_temp:.1f}°C")

            # 調整欄寬
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 16

            # 生成帶標記的熱力圖影像並暫存為 jpg，固定放在 E1
            imageA_input = cv2.cvtColor(np.array(self.imageA), cv2.COLOR_RGB2BGR)
            imageA_output = draw_numpy_image_item(imageA_input, self.mark_rect_A)
            temp_dir = tempfile.mkdtemp()
            temp_img_path = os.path.join(temp_dir, "thermal_report.jpg")
            Image.fromarray(cv2.cvtColor(imageA_output, cv2.COLOR_BGR2RGB)).save(temp_img_path, quality=100)

            img = XlImage(temp_img_path)
            # 限制影像寬度，保持比例
            max_width = 800
            if img.width > max_width:
                ratio = max_width / img.width
                img.width = max_width
                img.height = int(img.height * ratio)
            ws.add_image(img, "E1")

            wb.save(report_path)

            # 清理暫存檔案
            try:
                os.remove(temp_img_path)
                os.rmdir(temp_dir)
            except Exception:
                pass

            show_toast(
                title='寫入成功',
                message=f"已將數據寫入測試報告 Sheet「{final_name}」",
                duration=5000,
                toast_type='success'
            )
        except PermissionError:
            show_toast(
                title='寫入失敗',
                message="測試報告檔案已被其他程式開啟，請先關閉後再試",
                duration=5000,
                toast_type='error'
            )
        except Exception as e:
            print(f"寫入測試報告時出錯: {e}")
            import traceback
            traceback.print_exc()
            show_toast(
                title='寫入失敗',
                message=f"寫入測試報告時發生錯誤: {e}",
                duration=5000,
                toast_type='error'
            )

    def open_settings_dialog(self):
        # 使用单例模式，只创建一个SettingDialog实例
        if self.setting_dialog is None:
            self.setting_dialog = SettingDialog(self.root, self.settings_button, None)
            # 将 update_content 方法附加到 root 上，以便 SettingDialog 可以调用
            self.root.update_content = self.update_content
        self.setting_dialog.open()

    def update_content(self):
        """設定對話框確認後的回調函數。

        當使用者在設定對話框中修改字體大小或顏色後，
        此方法會被調用以立即更新 canvasA 和 canvasB 上的顯示。
        """
        print("update_content 被調用 - 重新繪製 canvas")

        # 如果有溫度標記數據，重新繪製 canvas
        if hasattr(self, 'mark_rect_A') and len(self.mark_rect_A) > 0:
            print(f"重新繪製 {len(self.mark_rect_A)} 個溫度標記")
            self.update_images()
        else:
            print("沒有溫度標記數據，跳過重新繪製")

    def load_local_image(self, index):
        img_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.png;*.jpeg")])
        if not img_path:
            return
        self.set_image(img_path, index)
    def set_image(self, path = Constants.imageA_default_path, index = 0):
        try:
            if not os.path.exists(path):
                print(f"图片文件不存在: {path}")
                return
                
            if index == 0:
                self.imageA = Image.open(path)
                self.mark_rect_A = []
                # 切换热力图后，清空对应打点，避免误判有点位
                self.points_A = []
                print(f"成功加载热力图: {path}")
                # 比對熱力圖與溫度數據的解析度
                self._validate_image_temp_dimensions()
            elif index == 1:
                self.imageB = Image.open(path)
                # 直接使用原图，不再强制缩放
                print(f"Layout图像尺寸: {self.imageB.size}")
                self.mark_rect_B = []
                # 切换Layout图后，清空对应打点，避免误判有点位
                self.points_B = []
                print(f"成功加载Layout图: {path}")

            if self.imageB and self.imageA:
                print(f"图像尺寸 - 热力图: {self.imageA.width}x{self.imageA.height}, Layout图: {self.imageB.width}x{self.imageB.height}")
            
            # 无论是否两个图片都加载完成，都尝试更新显示
            self.update_images()

            # 尝试自动加载与当前文件组合对应的打点JSON
            self.load_points()
            # 根据是否存在打点数据，更新对齐按钮可见性
            if hasattr(self, 'update_align_buttons_visibility'):
                self.update_align_buttons_visibility()
        except Exception as e:
            print(f"加载图片时出错: {e}")
    def _show_blended_window(self, title, blended_bgr):
        """在 Toplevel 視窗中顯示重疊圖像，縮放/全螢幕時保持熱力圖原圖長寬比"""
        blended_rgb = cv2.cvtColor(blended_bgr, cv2.COLOR_BGR2RGB)
        pil_img = Image.fromarray(blended_rgb)

        top = tk.Toplevel(self.root)
        top.title(title)
        top.configure(bg='black')

        canvas = tk.Canvas(top, bg='black', highlightthickness=0)
        canvas.pack(fill=tk.BOTH, expand=True)

        # 保留引用避免 GC
        top._pil_img = pil_img
        top._tk_img = None

        def _on_resize(event):
            cw, ch = event.width, event.height
            if cw < 2 or ch < 2:
                return
            iw, ih = pil_img.size
            scale = min(cw / iw, ch / ih)
            new_w = max(1, int(iw * scale))
            new_h = max(1, int(ih * scale))
            resized = pil_img.resize((new_w, new_h), Image.LANCZOS)
            top._tk_img = ImageTk.PhotoImage(resized)
            canvas.delete("all")
            canvas.create_image(cw // 2, ch // 2, image=top._tk_img, anchor='center')

        canvas.bind('<Configure>', _on_resize)

        # 初始視窗大小：螢幕 80% 內等比縮放
        screen_w = self.root.winfo_screenwidth()
        screen_h = self.root.winfo_screenheight()
        iw, ih = pil_img.size
        init_scale = min((screen_w * 0.8) / iw, (screen_h * 0.8) / ih, 1.0)
        init_w = max(400, int(iw * init_scale))
        init_h = max(300, int(ih * init_scale))
        top.geometry(f"{init_w}x{init_h}")

    def margin_before(self):
        try:
            # 检查图像是否存在
            if not hasattr(self, 'resized_imageA') or not hasattr(self, 'resized_imageB') or \
               self.resized_imageA is None or self.resized_imageB is None:
                print("警告：图像数据不存在，无法进行图像混合")
                return

            # 将 Pillow 图像对象转换为 NumPy 数组
            imageB_np = np.array(self.resized_imageB)
            imageA_np = np.array(self.resized_imageA)

            print(f"margin_before - 原始图像形状 - imageA: {imageA_np.shape}, imageB: {imageB_np.shape}")

            # 检查图像尺寸是否匹配
            if imageA_np.shape != imageB_np.shape:
                print(f"警告：图像尺寸不匹配 - imageA: {imageA_np.shape}, imageB: {imageB_np.shape}")
                # 将imageB调整到与imageA相同的尺寸
                imageB_np = cv2.resize(imageB_np, (imageA_np.shape[1], imageA_np.shape[0]))
                print(f"调整后 - imageB: {imageB_np.shape}")

            # 如果是 RGB 图像，OpenCV 默认处理 BGR 格式，所以需要转换颜色顺序
            if imageB_np.ndim == 3:  # 这是 RGB 图像
                imageB_np = cv2.cvtColor(imageB_np, cv2.COLOR_RGB2BGR)
                imageA_np = cv2.cvtColor(imageA_np, cv2.COLOR_RGB2BGR)
                print(f"颜色转换后 - imageA: {imageA_np.shape}, imageB: {imageB_np.shape}")

            # 最终检查
            if imageA_np.shape != imageB_np.shape:
                print(f"错误：无法使两个图像尺寸匹配 - imageA: {imageA_np.shape}, imageB: {imageB_np.shape}")
                return

            print(f"开始图像混合 - imageA: {imageA_np.shape}, imageB: {imageB_np.shape}")
            blended = cv2.addWeighted(imageB_np, 0.33, imageA_np, 0.66, 0)
            self._show_blended_window('对齐前图像', blended)
            print("margin_before 图像混合完成")

        except Exception as e:
            print(f"margin_before 方法出错: {e}")
            import traceback
            traceback.print_exc()

    def margin_after(self):
        try:
            # 检查point_transformer是否存在
            if self.point_transformer is None:
                print("警告：point_transformer为None，无法进行图像对齐")
                return

            # 检查图像是否存在
            if not hasattr(self, 'resized_imageA') or not hasattr(self, 'resized_imageB') or \
               self.resized_imageA is None or self.resized_imageB is None:
                print("警告：图像数据不存在，无法进行图像对齐")
                return

            bW, bH = self.resized_imageB.size
            aW, aH = self.resized_imageA.size

            print(f"图像尺寸 - imageA: {aW}x{aH}, imageB: {bW}x{bH}")

            # 将 Pillow 图像对象转换为 NumPy 数组
            imageB_np = np.array(self.resized_imageB)
            imageA_np = np.array(self.resized_imageA)

            print(f"NumPy数组形状 - imageA: {imageA_np.shape}, imageB: {imageB_np.shape}")

            # 如果是 RGB 图像，OpenCV 默认处理 BGR 格式，所以需要转换颜色顺序
            if imageB_np.ndim == 3:  # 这是 RGB 图像
                imageB_np = cv2.cvtColor(imageB_np, cv2.COLOR_RGB2BGR)
                imageA_np = cv2.cvtColor(imageA_np, cv2.COLOR_RGB2BGR)
                print(f"颜色转换后 - imageA: {imageA_np.shape}, imageB: {imageB_np.shape}")

            # 获取原始坐标系下的 B->A 变换矩阵
            M_ori = self.point_transformer.get_B2A_matrix()
            M_ori = np.asarray(M_ori)
            print(f"原始坐标变换矩阵形状: {M_ori.shape}")

            # 将原始坐标变换矩阵换算到当前显示尺寸（resized）
            sA = float(self.imageA_scale)
            sB = float(self.imageB_scale)
            if M_ori.shape == (2, 3):
                # Affine: pA_ori = A * pB_ori + t
                A = M_ori[:, :2]
                t = M_ori[:, 2:3]
                A_disp = (sA / sB) * A
                t_disp = sA * t
                M_disp = np.hstack([A_disp, t_disp]).astype(np.float32)
                aligned_imageB = cv2.warpAffine(imageB_np, M_disp, (aW, aH))
            elif M_ori.shape == (3, 3):
                # Homography: H_disp = S_A * H_ori * S_B^{-1}
                S_A = np.array([[sA, 0, 0], [0, sA, 0], [0, 0, 1]], dtype=np.float32)
                S_B_inv = np.array([[1.0 / sB, 0, 0], [0, 1.0 / sB, 0], [0, 0, 1]], dtype=np.float32)
                H_disp = (S_A @ M_ori @ S_B_inv).astype(np.float32)
                aligned_imageB = cv2.warpPerspective(imageB_np, H_disp, (aW, aH))
            else:
                print(f"未知的变换矩阵尺寸: {M_ori.shape}")
                return
            print(f"对齐后图像形状: {aligned_imageB.shape}")

            # 检查对齐后的图像尺寸是否与imageA匹配
            if aligned_imageB.shape != imageA_np.shape:
                print(f"警告：对齐后图像尺寸不匹配 - aligned_imageB: {aligned_imageB.shape}, imageA: {imageA_np.shape}")
                # 如果仍然不匹配，调整aligned_imageB的尺寸
                aligned_imageB = cv2.resize(aligned_imageB, (imageA_np.shape[1], imageA_np.shape[0]))
                print(f"调整后图像形状: {aligned_imageB.shape}")

            # 最终检查两个图像的形状是否完全匹配
            if aligned_imageB.shape != imageA_np.shape:
                print(f"错误：无法使两个图像尺寸匹配 - aligned_imageB: {aligned_imageB.shape}, imageA: {imageA_np.shape}")
                return

            print(f"开始图像混合 - aligned_imageB: {aligned_imageB.shape}, imageA: {imageA_np.shape}")
            blended = cv2.addWeighted(aligned_imageB, 0.33, imageA_np, 0.66, 0)
            self._show_blended_window('对齐后图像', blended)
            print("图像混合完成")

        except Exception as e:
            print(f"margin_after 方法出错: {e}")
            import traceback
            traceback.print_exc()

    def on_double_click(self, event):
        if len(self.mark_rect_A) > 0:
            # 检查EditorCanvas是否已经存在且可见
            if self.editor_canvas is not None and hasattr(self.editor_canvas, 'dialog') and self.editor_canvas.dialog.winfo_exists():
                # 如果EditorCanvas已存在，将其提到前台
                self.editor_canvas.dialog.lift()
                self.editor_canvas.dialog.focus_force()
                return
            
            # 创建新的EditorCanvas实例
            # 传递self作为parent，这样EditorCanvas可以访问到layout_data、point_transformer等属性
            self.editor_canvas = EditorCanvas(
                self, self.imageA, self.mark_rect_A, self.on_close_editor,
                self.current_temp_file_path,
                origin_mark_rect=getattr(self, 'origin_mark_rect_A', None)
            )
    
    def init_UI_flow(self, root):
        # 创建顶部按钮区域
        self.top_buttons_frame = tk.Frame(root, borderwidth=1, relief=tk.SUNKEN, bg=UIStyle.VERY_LIGHT_BLUE)
        self.top_buttons_frame.grid(row=0, column=0, columnspan=3, sticky="ew")
        self.top_buttons_frame.pack_propagate(False)  # 防止自动调整大小
        # 顶部按钮按钮
        self.folder_control_button = tk.Button(self.top_buttons_frame, text="隐藏文件夹Tab", command=self.toggle_folder_panel, 
                                             width=16, bg=UIStyle.SUCCESS_GREEN, fg=UIStyle.WHITE, 
                                             relief=UIStyle.BUTTON_RELIEF, borderwidth=UIStyle.BUTTON_BORDER_WIDTH,
                                             font=UIStyle.BUTTON_FONT)
        self.align_button = tk.Button(self.top_buttons_frame, text="多点对齐", command=self.start_point_mark,
                                     width=16, bg=UIStyle.WARNING_ORANGE, fg=UIStyle.WHITE,
                                     relief=UIStyle.BUTTON_RELIEF, borderwidth=UIStyle.BUTTON_BORDER_WIDTH,
                                     font=UIStyle.BUTTON_FONT)
        self.rect_align_button = tk.Button(self.top_buttons_frame, text="矩形对齐", command=self.start_rect_align,
                                          width=16, bg=UIStyle.WARNING_ORANGE, fg=UIStyle.WHITE,
                                          relief=UIStyle.BUTTON_RELIEF, borderwidth=UIStyle.BUTTON_BORDER_WIDTH,
                                          font=UIStyle.BUTTON_FONT)
        def debug_open_template_dialog():
            print("温度过滤按钮被点击！")
            self.open_template_dialog()
        
        self.template_filter_button = tk.Button(self.top_buttons_frame, text="温度过滤", command=debug_open_template_dialog, 
                                             width=10, bg=UIStyle.PRIMARY_BLUE, fg=UIStyle.WHITE, 
                                             relief=UIStyle.BUTTON_RELIEF, borderwidth=UIStyle.BUTTON_BORDER_WIDTH,
                                             font=UIStyle.BUTTON_FONT)
        self.export_button = tk.Button(self.top_buttons_frame, text="导出", command=self.export_excel,
                                     width=10, bg=UIStyle.DARK_BLUE, fg=UIStyle.WHITE,
                                     relief=UIStyle.BUTTON_RELIEF, borderwidth=UIStyle.BUTTON_BORDER_WIDTH,
                                     font=UIStyle.BUTTON_FONT)
        self.test_report_button = tk.Button(self.top_buttons_frame, text="寫入測試報告", command=self.write_test_report,
                                           width=14, bg=UIStyle.DARK_BLUE, fg=UIStyle.WHITE,
                                           relief=UIStyle.BUTTON_RELIEF, borderwidth=UIStyle.BUTTON_BORDER_WIDTH,
                                           font=UIStyle.BUTTON_FONT, state=tk.DISABLED)
        self.settings_button = tk.Button(self.top_buttons_frame, text="设置", command=self.open_settings_dialog,
                                       width=10, bg=UIStyle.GRAY, fg=UIStyle.WHITE,
                                       relief=UIStyle.BUTTON_RELIEF, borderwidth=UIStyle.BUTTON_BORDER_WIDTH,
                                       font=UIStyle.BUTTON_FONT)

        self.folder_control_button.grid(row=0, column=0, padx=8, pady=16)
        self.align_button.grid(row=0, column=1, padx=8, pady=16)
        self.rect_align_button.grid(row=0, column=2, padx=8, pady=16)
        self.template_filter_button.grid(row=0, column=3, padx=8, pady=16)
        self.export_button.grid(row=0, column=4, padx=8, pady=16)
        self.test_report_button.grid(row=0, column=5, padx=8, pady=16)
        self.settings_button.grid(row=0, column=6, padx=8, pady=16)

        # 创建文件夹选择区域（固定宽度220像素，不可扩展）
        self.folder_container = tk.Frame(root, bg=UIStyle.VERY_LIGHT_BLUE, relief=tk.SUNKEN, bd=1, width=230)
        self.folder_container.grid(row=1, column=0, sticky="ns", padx=5, pady=5)
        self.folder_container.pack_propagate(False)  # 防止内容改变容器大小
        
        self.folder_frame = tk.Frame(self.folder_container, borderwidth=0, relief=tk.FLAT, bg=UIStyle.WHITE)
        self.folder_frame.pack(fill=tk.BOTH, expand=True, pady=10, padx=10)
        
        # 合并的文件夹信息行
        self.folder_info_frame = tk.Frame(self.folder_frame, bg=UIStyle.WHITE)
        self.folder_info_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 左侧：当前文件夹信息（80%宽度，绿色背景）
        self.folder_path_frame = tk.Frame(self.folder_info_frame, bg=UIStyle.SUCCESS_GREEN, relief=tk.FLAT, bd=0)
        self.folder_path_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        self.folder_path_label = tk.Label(self.folder_path_frame, text="当前文件夹：未选择", 
                                        bg=UIStyle.SUCCESS_GREEN, fg=UIStyle.WHITE, relief=tk.FLAT, bd=0,
                                        font=UIStyle.BUTTON_FONT, anchor=tk.W, wraplength=160, justify=tk.LEFT)
        self.folder_path_label.pack(pady=8, padx=10, fill=tk.BOTH, expand=True)
        
        # 右侧：选择文件夹按钮（20%宽度，无背景色，与顶部按钮高度一致）
        self.folder_button = tk.Button(self.folder_info_frame, text="📂", command=self.select_folder, 
                                     width=3, bg=UIStyle.WHITE, fg=UIStyle.SUCCESS_GREEN, 
                                     relief=tk.FLAT, borderwidth=0,
                                     font=("Arial", 16))
        self.folder_button.pack(side=tk.RIGHT, padx=(5, 0), fill=tk.Y)
        
        # 文件分类树形视图
        self.folder_tree = ttk.Treeview(self.folder_frame, height=10, show="tree")
        
        # 配置Treeview样式，支持加粗标记
        style = ttk.Style()
        style.configure("Treeview", 
                       foreground=UIStyle.BLACK, 
                       background=UIStyle.WHITE,
                       fieldbackground=UIStyle.WHITE,
                       borderwidth=1,
                       relief="solid")
        style.configure("Treeview.Item", 
                       foreground=UIStyle.BLACK,
                       background=UIStyle.WHITE)
        
        # 创建加粗标记的标签样式
        style.configure("Bold.Treeview.Item", 
                       font=UIStyle.BUTTON_FONT,
                       foreground=UIStyle.DARK_BLUE,
                       background=UIStyle.VERY_LIGHT_BLUE)
        
        # 配置选中项样式
        style.map("Treeview", 
                 background=[('selected', UIStyle.LIGHT_BLUE)],
                 foreground=[('selected', UIStyle.WHITE)])
        
        self.folder_tree.pack(pady=10, padx=5, fill=tk.BOTH, expand=True)
        
        # 配置Treeview的标签样式
        self.folder_tree.tag_configure("bold", 
                                      font=UIStyle.BUTTON_FONT,
                                      foreground=UIStyle.DARK_BLUE,
                                      background=UIStyle.VERY_LIGHT_BLUE)
        
        # 绑定单击事件
        self.folder_tree.bind("<Button-1>", self.on_file_click)

        # 綁定 Tooltip 事件
        self._tree_tooltip = None  # Tooltip Toplevel 視窗
        self._tree_tooltip_item = None  # 當前顯示 tooltip 的 item
        self.folder_tree.bind("<Motion>", self._on_tree_motion)
        self.folder_tree.bind("<Leave>", self._hide_tree_tooltip)
        
        # 去除滚动条，让内容占满整个区域
        # folder_scrollbar = ttk.Scrollbar(self.folder_frame, orient="vertical", command=self.folder_tree.yview)
        # folder_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        # self.folder_tree.configure(yscrollcommand=folder_scrollbar.set)

        #中间图片区域
        # self.imageA = Image.open(Constants.imageA_default_path)
        # self.imageB = Image.open(Constants.imageB_default_path)
       
        # 创建 Canvas，使用 grid 布局来控制横向排列
        self.canvasA = tk.Canvas(root, bg=UIStyle.WHITE, relief=tk.SUNKEN, bd=1, 
                                highlightthickness=0, highlightbackground=UIStyle.LIGHT_GRAY)
        self.canvasB = tk.Canvas(root, bg=UIStyle.WHITE, relief=tk.SUNKEN, bd=1,
                                highlightthickness=0, highlightbackground=UIStyle.LIGHT_GRAY)

        self.canvasA.grid(row=1, column=1, sticky="nsew")
        self.canvasB.grid(row=1, column=2, sticky="nsew")
        # 让 Grid 布局管理器将列的权重设置为1，使得画布可以在横向上均匀分配空间
        root.grid_rowconfigure(1, weight=1)
        root.grid_columnconfigure(0, weight=0)  # 文件夹区域不拉伸
        root.grid_columnconfigure(1, weight=1)  # 画布A拉伸
        root.grid_columnconfigure(2, weight=1)  # 画布B拉伸
        # 设定变量来存储图像引用，避免重复创建
        self.tk_imageA = None
        self.tk_imageB = None

        # 不再默认加载图片，等待用户选择文件夹
        # self.set_image(Constants.imageA_default_path, 0)
        # self.set_image(Constants.imageB_default_path, 1)

         # 创建下方按钮区域
        self.bottom_buttons_frame_A = tk.Frame(root, bg=UIStyle.VERY_LIGHT_BLUE, relief=tk.FLAT, bd=0)
        self.bottom_buttons_frame_A.grid(row=2, column=1, padx=8, pady=16)
        # self.point_imageA_button = tk.Button(self.bottom_buttons_frame_A, text="打点标记", command=self.point_mark_A, width=10)
        self.margin_before_button = tk.Button(self.bottom_buttons_frame_A, text="对齐前图像", command=self.margin_before, 
                                            width=10, bg=UIStyle.PRIMARY_BLUE, fg=UIStyle.WHITE, 
                                            relief=UIStyle.BUTTON_RELIEF, borderwidth=UIStyle.BUTTON_BORDER_WIDTH,
                                            font=UIStyle.BUTTON_FONT)
        self.margin_after_button = tk.Button(self.bottom_buttons_frame_A, text="对齐后图像", command=self.margin_after, 
                                           width=10, bg=UIStyle.PRIMARY_BLUE, fg=UIStyle.WHITE, 
                                           relief=UIStyle.BUTTON_RELIEF, borderwidth=UIStyle.BUTTON_BORDER_WIDTH,
                                           font=UIStyle.BUTTON_FONT)
        
        # 添加清除热力图对齐点按钮
        self.clear_heat_points_button = tk.Button(self.bottom_buttons_frame_A, text="清除热力图对齐点", 
                                                command=self.clear_heat_points, width=15, bg=UIStyle.PRIMARY_BLUE, fg=UIStyle.WHITE, 
                                                relief=UIStyle.BUTTON_RELIEF, borderwidth=UIStyle.BUTTON_BORDER_WIDTH,
                                                font=UIStyle.BUTTON_FONT)
       
        self.bottom_buttons_frame_A.pack_propagate(False)  # 不允许frame自动调整大小
        # 初始根据是否存在打点JSON控制显示
        self.margin_before_button.grid_forget()
        self.margin_after_button.grid_forget()
        self.update_align_buttons_visibility()
        # 初始隐藏清除按钮
        self.clear_heat_points_button.grid_forget()

        self.bottom_buttons_frame_B = tk.Frame(root, bg=UIStyle.VERY_LIGHT_BLUE, relief=tk.FLAT, bd=0)
        self.bottom_buttons_frame_B.grid(row=2, column=2)
        # self.point_imageB_button = tk.Button(self.bottom_buttons_frame_B, text="打点标记", command=self.point_mark_B, width=10)
        
        # 添加清除Layout图对齐点按钮
        self.clear_layout_points_button = tk.Button(self.bottom_buttons_frame_B, text="清除Layout图对齐点", 
                                               command=self.clear_layout_points, width=15, bg=UIStyle.PRIMARY_BLUE, fg=UIStyle.WHITE, 
                                               relief=UIStyle.BUTTON_RELIEF, borderwidth=UIStyle.BUTTON_BORDER_WIDTH,
                                               font=UIStyle.BUTTON_FONT)
        
        self.bottom_buttons_frame_B.pack_propagate(False)  # 不允许frame自动调整大小
        # 初始隐藏清除按钮
        self.clear_layout_points_button.grid_forget()
        # self.point_imageB_button.grid(row=0, column=1, padx=5)    

        # self.imgScalePCB = self.imageB.width / self.imageA.width
        # print("self.imgScalePCB -> ", self.imgScalePCB, self.imageB.width, self.imageA.width)
        # self.root.after(100, self.init_point_transformer)

        self.canvasA.bind("<Double-Button-1>", self.on_double_click) # 左键
    def has_points_json(self):
        try:
            if not self.current_folder_path:
                return False
            points_dir = os.path.join(self.current_folder_path, "points")
            heat_filename = self.current_files.get("heat", "")
            layout_filename = self.current_files.get("layout", "")
            if not heat_filename or not layout_filename:
                # 如果文件名未就绪，但内存中已有足够的点位，也认为可显示
                return len(self.points_A) >= 3 and len(self.points_B) >= 3
            heat_name = os.path.splitext(heat_filename)[0]
            layout_name = os.path.splitext(layout_filename)[0]
            json_points_path = os.path.join(points_dir, f"{heat_name}_{layout_name}.json")
            return os.path.exists(json_points_path) or (len(self.points_A) >= 3 and len(self.points_B) >= 3)
        except Exception:
            return False
    def update_align_buttons_visibility(self):
        if self.has_points_json():
            self.margin_before_button.grid(row=0, column=0, padx=5)
            self.margin_after_button.grid(row=0, column=1, padx=5)
        else:
            self.margin_before_button.grid_forget()
            self.margin_after_button.grid_forget()
    def load_local_data(self):
        """加载本地数据，现在从选择的文件夹中加载"""
        if self.current_folder_path:
            # 如果已经选择了文件夹，从文件夹中加载数据
            self.scan_folder_files()
            # 扫描后刷新按钮显示状态
            self.update_align_buttons_visibility()
        else:
            # 如果没有选择文件夹，加载默认数据
            self.load_points()

    def to_numpy_image(self, image):
        image_np = np.array(image)
        # 如果是 RGB 图像，OpenCV 默认处理 BGR 格式，所以需要转换颜色顺序
        if image_np.ndim == 3:  # 这是 RGB 图像
            image_np = cv2.cvtColor(image_np, cv2.COLOR_RGB2BGR)
        return image_np
    def save_log_file(self):
        # 获取当前时间并格式化为字符串
        current_year = datetime.now().strftime("%Y")
        current_time = datetime.now().strftime("%m-%d %H:%M")

        self.edit_log["export_time"][1] = current_time

        if not os.path.exists("logs"):
            os.makedirs("logs")  # 创建多层目录

        # 生成 CSV 文件名
        csv_filename = "logs/" + f"{current_year}.csv"

        # 检查文件是否存在
        file_exists = os.path.exists(csv_filename)

        # 打开文件并写入 CSV
        with open(csv_filename, mode='a', newline='', encoding='utf-8-sig') as file:
            writer = csv.writer(file)

            # 如果文件不存在，则写入头部
            if not file_exists:
                header = ["export_time", "origin_mark", "final_mark", "add_new_mark", "delete_origin_mark", "modify_origin_mark"]
                writer.writerow(header)  # 写入头部
                header_values = ["生成时间", "自动生成外框数量", "最终导出外框数量", "新增外框数量（手动增加导出时没有被删除）", "删除外框数量（自动生成的外框被删除）", "调整外框数量（自动生成的外框被调整)"]
                writer.writerow(header_values)  # 写入描述行

            # 遍历字典并写入每一行数值
            row = []
            for key, value in self.edit_log.items():
                # 如果是 'modify_origin_mark' 并且 value[1] 是 set，则取 set 的长度
                target_value = None
                if isinstance(value[1], set):
                    target_value = len(value[1])  # 取 set 的大小
                else:
                    target_value = value[1]

                row.append(target_value)

            writer.writerow(row)  # 写入数据行

        print(f"CSV 文件已保存为 {csv_filename}")

    # ==================== 矩形對齊功能 ====================

    def _enter_rect_fullscreen(self):
        """進入矩形對齊滿版模式：隱藏 canvasB，讓熱力圖佔滿"""
        self.canvasB.grid_forget()
        self.bottom_buttons_frame_B.grid_forget()
        self.root.grid_columnconfigure(2, weight=0)
        self.root.after(100, self.update_images)

    def _exit_rect_fullscreen(self):
        """退出矩形對齊滿版模式：恢復 canvasB"""
        self.canvasB.grid(row=1, column=2, sticky="nsew")
        self.root.grid_columnconfigure(2, weight=1)
        self.root.after(100, self.update_images)

    def start_rect_align(self):
        """矩形對齊按鈕的主要處理邏輯（進入模式 / 確認對齊）"""
        if not self.is_rect_aligning:
            # ---- 進入矩形對齊模式 ----
            if not self.imageA or not self.imageB:
                show_toast(title='缺少圖片', message='請先載入熱力圖和Layout圖', duration=3000, toast_type='warning')
                return

            self.is_rect_aligning = True
            # 保留已有的 rect_corners（來自上次對齊或 JSON），進入後直接顯示供微調
            self.rect_drag_start = None
            self.rect_dragging_corner = None

            self._enter_rect_fullscreen()
            self._bind_rect_events()

            # 禁用多點對齊按鈕
            self.align_button.config(state=tk.DISABLED)
            self.rect_align_button.config(text="确认矩形对齐")

        else:
            # ---- 確認矩形對齊 ----
            if self.rect_corners is None or len(self.rect_corners) != 4:
                show_toast(title='尚未框選', message='請先在熱力圖上拖曳矩形框選 PCBA 區域', duration=3000, toast_type='warning')
                return

            # Layout 圖的 4 個角落
            bW, bH = self.imageB.size
            layout_corners = [(0, 0), (bW, 0), (bW, bH), (0, bH)]

            try:
                transformer = PointTransformer(self.rect_corners, layout_corners, matched=True)
            except Exception as e:
                show_toast(title='对齐失败', message=f'矩形对齐异常：{e}', duration=5000, toast_type='error')
                return

            # 成功 —— 儲存結果
            self.point_transformer = transformer
            self.alignment_type = 'rect'

            # 將 rect_corners 也存入 points_A / points_B 以便 JSON 序列化
            self.points_A = list(self.rect_corners)
            self.points_B = layout_corners

            self.save_points_json()
            self.update_align_buttons_visibility()

            # 清理模式
            self._unbind_rect_events()
            self._clear_rect_overlay()
            self.is_rect_aligning = False

            self._exit_rect_fullscreen()

            # 恢復按鈕
            self.align_button.config(state=tk.NORMAL)
            self.rect_align_button.config(text="矩形对齐")

            # 若之前有標記框，清除（重新對齊了）
            if self.pont_marked:
                self.mark_rect_A = []
                self.mark_rect_B = []
                self.pont_marked = False

            show_toast(title='矩形对齐成功', message='已完成矩形对齐，可使用温度过滤', duration=3000, toast_type='success')

    # ---- 矩形拖曳事件綁定 / 解綁 ----

    def _bind_rect_events(self):
        self.canvasA.bind("<ButtonPress-1>", self._on_rect_press)
        self.canvasA.bind("<B1-Motion>", self._on_rect_drag)
        self.canvasA.bind("<ButtonRelease-1>", self._on_rect_release)

    def _unbind_rect_events(self):
        self.canvasA.unbind("<ButtonPress-1>")
        self.canvasA.unbind("<B1-Motion>")
        self.canvasA.unbind("<ButtonRelease-1>")

    # ---- 座標轉換 ----

    def _canvas_to_original(self, cx, cy):
        """canvas 座標 → 原始圖像座標"""
        offX, offY = getattr(self, 'canvasA_offset', (0, 0))
        scale = getattr(self, 'imageA_scale', 1.0)
        if scale == 0:
            scale = 1.0
        ox = (cx - offX) / scale
        oy = (cy - offY) / scale
        return (ox, oy)

    def _original_to_canvas(self, ox, oy):
        """原始圖像座標 → canvas 座標"""
        offX, offY = getattr(self, 'canvasA_offset', (0, 0))
        scale = getattr(self, 'imageA_scale', 1.0)
        cx = ox * scale + offX
        cy = oy * scale + offY
        return (cx, cy)

    def _find_corner_at(self, cx, cy, threshold=12):
        """在 canvas 座標附近尋找已有角點，回傳角索引 0-3 或 None"""
        if self.rect_corners is None:
            return None
        for i, (ox, oy) in enumerate(self.rect_corners):
            ccx, ccy = self._original_to_canvas(ox, oy)
            if abs(cx - ccx) <= threshold and abs(cy - ccy) <= threshold:
                return i
        return None

    # ---- 拖曳事件處理 ----

    def _on_rect_press(self, event):
        cx, cy = event.x, event.y
        # 先嘗試抓取已有角點
        corner_idx = self._find_corner_at(cx, cy)
        if corner_idx is not None:
            self.rect_dragging_corner = corner_idx
            self.rect_drag_start = None
            return
        # 否則開始新矩形拖曳
        self.rect_drag_start = (cx, cy)
        self.rect_dragging_corner = None

    def _on_rect_drag(self, event):
        cx, cy = event.x, event.y
        if self.rect_dragging_corner is not None:
            # 移動單一角點
            ox, oy = self._canvas_to_original(cx, cy)
            self.rect_corners[self.rect_dragging_corner] = (ox, oy)
            self._draw_rect_overlay()
        elif self.rect_drag_start is not None:
            # 畫新矩形
            sx, sy = self.rect_drag_start
            tl = self._canvas_to_original(min(sx, cx), min(sy, cy))
            tr = self._canvas_to_original(max(sx, cx), min(sy, cy))
            br = self._canvas_to_original(max(sx, cx), max(sy, cy))
            bl = self._canvas_to_original(min(sx, cx), max(sy, cy))
            self.rect_corners = [tl, tr, br, bl]
            self._draw_rect_overlay()

    def _on_rect_release(self, event):
        self.rect_dragging_corner = None
        self.rect_drag_start = None

    # ---- 矩形 Overlay 繪製 ----

    def _clear_rect_overlay(self):
        for item_id in self.rect_canvas_ids:
            self.canvasA.delete(item_id)
        self.rect_canvas_ids = []

    def _draw_rect_overlay(self):
        """繪製矩形四邊形 overlay（4 綠線 + 彩色角手柄）"""
        self._clear_rect_overlay()
        if self.rect_corners is None or len(self.rect_corners) != 4:
            return

        canvas_pts = [self._original_to_canvas(ox, oy) for (ox, oy) in self.rect_corners]

        # 4 條邊線（綠色）
        for i in range(4):
            x1, y1 = canvas_pts[i]
            x2, y2 = canvas_pts[(i + 1) % 4]
            lid = self.canvasA.create_line(x1, y1, x2, y2, fill="#00FF00", width=2, tags="rect_overlay")
            self.rect_canvas_ids.append(lid)

        # 4 個角手柄（TL=紅, TR=綠, BR=藍, BL=黃）
        colors = ["#FF0000", "#00CC00", "#0066FF", "#FFCC00"]
        labels = ["TL", "TR", "BR", "BL"]
        r = 6
        for i, (cx, cy) in enumerate(canvas_pts):
            oid = self.canvasA.create_oval(cx - r, cy - r, cx + r, cy + r,
                                           fill=colors[i], outline="white", width=2, tags="rect_overlay")
            self.rect_canvas_ids.append(oid)
            tid = self.canvasA.create_text(cx, cy - r - 8, text=labels[i],
                                           fill=colors[i], font=("Arial", 9, "bold"), tags="rect_overlay")
            self.rect_canvas_ids.append(tid)

    def toggle_folder_panel(self):
        """切换文件夹面板的可见性"""
        if self.folder_container.winfo_ismapped():
            # 隐藏整个文件夹容器
            self.folder_container.grid_forget()
            # 更新按钮文字
            self.folder_control_button.config(text="显示文件夹Tab")
            
            # 使用延迟更新机制，避免卡顿
            self.root.after(100, self._optimize_layout_after_hide)
        else:
            # 显示文件夹容器
            self.folder_container.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
            # 更新按钮文字
            self.folder_control_button.config(text="隐藏文件夹Tab")
            
            # 使用延迟更新机制，避免卡顿
            self.root.after(100, self._optimize_layout_after_show)
    
    def _optimize_layout_after_hide(self):
        """隐藏文件夹后的布局优化"""
        # 让右边的图片占满空间，使用更平滑的权重变化
        self.root.grid_columnconfigure(0, weight=0)  # 文件夹区域不拉伸
        self.root.grid_columnconfigure(1, weight=1)  # 画布A拉伸
        self.root.grid_columnconfigure(2, weight=1)  # 画布B拉伸
        
        # 延迟更新图片，避免卡顿
        if hasattr(self, 'imageA') and hasattr(self, 'imageB') and self.imageA and self.imageB:
            self.root.after(200, self._delayed_update_images)
    
    def _optimize_layout_after_show(self):
        """显示文件夹后的布局优化"""
        # 恢复原来的列权重设置
        self.root.grid_columnconfigure(0, weight=0)  # 文件夹区域不拉伸
        self.root.grid_columnconfigure(1, weight=1)  # 画布A拉伸
        self.root.grid_columnconfigure(2, weight=1)  # 画布B拉伸
        
        # 延迟更新图片，避免卡顿
        if hasattr(self, 'imageA') and hasattr(self, 'imageB') and self.imageA and self.imageB:
            self.root.after(200, self._delayed_update_images)
    
    def _delayed_update_images(self):
        """延迟更新图片，避免卡顿"""
        try:
            # 检查画布是否已经准备好
            if (hasattr(self, 'canvasA') and hasattr(self, 'canvasB') and 
                self.canvasA.winfo_width() > 1 and self.canvasB.winfo_width() > 1):
                self.update_images()
        except Exception as e:
            print(f"延迟更新图片时出错: {e}")

def setup_logging():
    """设置日志系统，将print输出重定向到日志文件"""
    # 创建logs目录
    logs_dir = "logs"
    if not os.path.exists(logs_dir):
        os.makedirs(logs_dir)
    
    # 生成日志文件名：年_月_日.txt
    today = datetime.now()
    log_filename = f"{today.year}_{today.month:02d}_{today.day:02d}.txt"
    log_filepath = os.path.join(logs_dir, log_filename)
    
    # 创建日志文件并写入启动信息
    with open(log_filepath, 'a', encoding='utf-8') as log_file:
        log_file.write(f"\n{'='*50}\n")
        log_file.write(f"程序启动时间: {today.strftime('%Y-%m-%d %H:%M:%S')}\n")
        log_file.write(f"{'='*50}\n")
    
    # 重定向stdout到日志文件
    class LogWriter:
        def __init__(self, file):
            self.file = file
            self.terminal = sys.stdout
        
        def write(self, message):
            self.terminal.write(message)  # 同时输出到控制台
            self.file.write(message)      # 写入日志文件
            self.file.flush()             # 立即刷新到文件
        
        def flush(self):
            self.terminal.flush()
            self.file.flush()
    
    # 打开日志文件并重定向stdout
    log_file = open(log_filepath, 'a', encoding='utf-8')
    sys.stdout = LogWriter(log_file)
    
    print(f"日志系统已启动，日志文件: {log_filepath}")
    return log_filepath

def parse_arguments():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(description='Thermal温度点位自动识别系统')
    parser.add_argument('--log', action='store_true', help='启用日志记录功能')
    return parser.parse_args()

if __name__ == "__main__":
    # 解析命令行参数
    args = parse_arguments()
    
    # 如果指定了--log参数，设置日志系统
    log_filepath = None
    if args.log:
        log_filepath = setup_logging()
    
    root = tk.Tk()
    root.configure(bg=UIStyle.VERY_LIGHT_BLUE)
    root.title("Thermal温度点位自动识别系统")
    root.geometry("1400x900")
    app = ResizableImagesApp(root)
    
    # 添加程序退出时的配置保存
    def on_closing():
        if log_filepath:
            print(f"程序退出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        app.save_current_files_to_config()
        root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_closing)
    root.mainloop()
